#!/usr/bin/env python3
"""
FulôFiló — Bootstrap Excel master workbook
===========================================
Creates data/excel/FuloFilo_Master.xlsx with all required sheets and headers.
Includes one placeholder SKU so a first sync can succeed with zero sales.

Usage:
  uv run python scripts/bootstrap_excel_master.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
EXCEL_DIR = ROOT / "data" / "excel"
OUT = EXCEL_DIR / "FuloFilo_Master.xlsx"


def main() -> None:
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # --- Catalog ---
    ws = wb.create_sheet("Catalog", 0)
    cat_headers = ["sku", "full_name", "category", "unit_cost", "suggested_price", "min_stock", "reorder_qty"]
    ws.append(cat_headers)
    ws.append([1, "Produto Exemplo", "Geral", 10.0, 25.0, 5, 15])

    # --- Inventory ---
    ws = wb.create_sheet("Inventory", 1)
    ws.append(["sku", "product", "category", "current_stock", "min_stock", "reorder_qty", "supplier", "lead_time_days", "notes"])
    ws.append([1, "Produto Exemplo", "Geral", 0, 5, 15, "", 7, ""])

    # --- DailySales (empty body) ---
    ws = wb.create_sheet("DailySales", 2)
    ws.append(["Date", "sku", "Product", "Quantity", "Unit_Price", "Total", "Payment_Method", "Source"])

    # --- Cashflow ---
    ws = wb.create_sheet("Cashflow", 3)
    ws.append(["Date", "Type", "Category", "Description", "Amount", "Payment_Method"])

    # --- CategoryOverrides ---
    ws = wb.create_sheet("CategoryOverrides", 4)
    ws.append(["sku", "category", "subcategory", "confidence"])

    # --- Meta ---
    ws = wb.create_sheet("Meta", 5)
    ws.append(["key", "value"])
    ws.append(["schema_version", "1"])
    ws.append(["workbook", "FuloFilo_Master"])

    wb.save(OUT)
    print(f"✅ Wrote {OUT}")
    print("   Next: uv run python scripts/sync_excel.py")


if __name__ == "__main__":
    main()
