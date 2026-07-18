#!/usr/bin/env python3
"""
FulôFiló — Sales Drop Watcher v3
==================================
Drop a file in data/incoming/ → full pipeline runs automatically.

Operational data flow (POS automation lane — coexists with Excel-canonical docs):
  1. Scan data/incoming/ for item-sales-summary-*.csv files
  2. Validate filename pattern + internal column signature
  3. Fuzzy product-name check — catch POS name drift before ingest
  4. Duplicate guard — warn if period already ingested
  5. Run etl/ingest.py → updates daily_sales / cashflow parquets + analytics
  6. Write back matching rows to FuloFilo_Master.xlsx (DailySales + Cashflow) under flock
  7. Run scripts/sync_excel.sh so derived Parquet/DuckDB match validated workbook schema
  8. Archive file → data/raw/
  9. git commit + push (optional) → notifications + dashboard

Manual lane remains: edit workbook → bash scripts/sync_excel.sh (see docs/documentation.md).

Usage:
    python scripts/sales_watcher.py              # process incoming/ once
    python scripts/sales_watcher.py --dry-run    # simulate, no writes
    python scripts/sales_watcher.py --daemon --interval 30   # LaunchAgent mode
"""

from __future__ import annotations

import argparse
import contextvars
import csv
import datetime
import difflib
import logging
import re
import shutil
import subprocess
import sys
import uuid
from pathlib import Path

from app.utils.workbook_lock import locked_workbook

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT         = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
INCOMING     = ROOT / "data" / "incoming"
ARCHIVE      = ROOT / "data" / "raw"
LOGS_DIR     = ROOT / "logs"
LOG_FILE     = LOGS_DIR / "saleswatch.log"
INGEST       = ROOT / "etl" / "ingest.py"
EXCEL_MASTER = ROOT / "data" / "excel" / "FuloFilo_Master.xlsx"
EXCEL_BACKUP = ROOT / "data" / "excel" / "backups"
PARQUET_DIR  = ROOT / "data" / "parquet"
DASHBOARD_URL = "https://autogio-fulofilo.streamlit.app/"

# ── File pattern ──────────────────────────────────────────────────────────────
FILENAME_RE = re.compile(
    r"^item-sales-summary-(\d{4}-\d{2}-\d{2})-(\d{4}-\d{2}-\d{2})\.csv$",
    re.IGNORECASE,
)

# Required columns (case-insensitive)
REQUIRED_COLS = {"sku", "itens vendidos", "vendas líquidas", "custo das mercadorias"}

# Fuzzy match threshold: 0.0 = anything, 1.0 = exact only
# 0.72 catches dash variants, accents, minor typos without false positives
FUZZY_CUTOFF = 0.72

# ── Logging (correlation id per process_file) ────────────────────────────────
_correlation_id: contextvars.ContextVar[str] = contextvars.ContextVar(
    "correlation_id", default="-"
)


class _CorrelationIdFilter(logging.Filter):
    def filter(self, record: logging.LogRecord) -> bool:
        record.correlation_id = _correlation_id.get()
        return True


LOGS_DIR.mkdir(parents=True, exist_ok=True)
_log_fmt = logging.Formatter(
    "%(asctime)s  %(correlation_id)s  %(levelname)-8s  %(message)s"
)
_fh = logging.FileHandler(LOG_FILE)
_sh = logging.StreamHandler(sys.stdout)
for _h in (_fh, _sh):
    _h.setFormatter(_log_fmt)
    _h.addFilter(_CorrelationIdFilter())

log = logging.getLogger("saleswatch")
log.handlers.clear()
log.setLevel(logging.INFO)
log.addHandler(_fh)
log.addHandler(_sh)
log.propagate = False


def _run_validate_sync_from_excel() -> tuple[bool, str]:
    """Re-run canonical sync so workbook remains SSoT for all derived artifacts."""
    script = ROOT / "scripts" / "sync_excel.sh"
    r = subprocess.run(
        ["bash", str(script)], cwd=str(ROOT), capture_output=True, text=True
    )
    parts = [p.strip() for p in (r.stdout, r.stderr) if p and p.strip()]
    return r.returncode == 0, "\n".join(parts)


# ══════════════════════════════════════════════════════════════════════════════
# VALIDATION HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _has_required_cols(path: Path) -> bool:
    try:
        with open(path, encoding="utf-8-sig") as f:
            header = next(csv.reader(f), [])
        cols = {c.lower().strip() for c in header}
        missing = REQUIRED_COLS - cols
        if missing:
            log.warning("  Column check failed — missing: %s", missing)
            return False
        return True
    except Exception as exc:
        log.error("  Could not read %s: %s", path.name, exc)
        return False


def _read_csv_summary(path: Path) -> tuple[list[str], float, float]:
    """
    Parse the CSV and return:
      (product_names, total_units, total_revenue)
    Uses only stdlib — safe to call before polars is involved.
    """
    names: list[str] = []
    units = 0.0
    revenue = 0.0
    try:
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = str(row.get("Item") or "").strip()
                if name:
                    names.append(name)
                try:
                    units += float(str(row.get("Itens vendidos") or "0").replace(",", "."))
                except ValueError:
                    pass
                try:
                    revenue += float(str(row.get("Vendas líquidas") or "0").replace(",", "."))
                except ValueError:
                    pass
    except Exception as exc:
        log.error("  Could not parse CSV summary: %s", exc)
    return names, units, revenue


def _load_known_product_names() -> list[str]:
    """Load full_name list from products.parquet. Returns [] if unavailable."""
    prod_path = PARQUET_DIR / "products.parquet"
    if not prod_path.exists():
        return []
    try:
        import polars as pl
        df = pl.read_parquet(prod_path)
        if "full_name" in df.columns:
            return df["full_name"].drop_nulls().to_list()
    except Exception:
        pass
    return []


def _normalize(name: str) -> str:
    """Normalize for fuzzy comparison: lowercase, collapse spaces, unify dashes."""
    name = name.lower().strip()
    # Unify all dash variants (en-dash –, em-dash —, hyphen -) to a single token
    name = re.sub(r"[–—\-]+", "—", name)
    # Collapse multiple spaces
    name = re.sub(r"\s+", " ", name)
    return name


def _fuzzy_validate_products(csv_names: list[str], known_names: list[str]) -> bool:
    """
    For each product name in the CSV, check for exact or close match in catalog.
    Logs WARN for fuzzy matches (name drift) and ERROR for no match at all.
    Returns True if all names are accounted for (exact or close enough).
    """
    if not known_names:
        log.info("  ℹ️  No catalog loaded — skipping product name validation")
        return True

    known_normalized = {_normalize(n): n for n in known_names}
    all_ok = True

    for csv_name in csv_names:
        norm = _normalize(csv_name)

        # ── Exact match (after normalization) ─────────────────────────────────
        if norm in known_normalized:
            continue

        # ── Fuzzy match ───────────────────────────────────────────────────────
        close = difflib.get_close_matches(norm, known_normalized.keys(), n=1, cutoff=FUZZY_CUTOFF)
        if close:
            catalog_name = known_normalized[close[0]]
            score = difflib.SequenceMatcher(None, norm, close[0]).ratio()
            log.warning(
                "  ⚠️  Name drift detected (%.0f%% match):", score * 100
            )
            log.warning("       CSV:     '%s'", csv_name)
            log.warning("       Catalog: '%s'  ← analytics will join on this", catalog_name)
            log.warning("       → Revenue will be attributed correctly via fuzzy join")
            # Not blocking — the fuzzy join note is informational
        else:
            # No match at all — this product won't appear in analytics
            log.error(
                "  ❌ No catalog match for: '%s'", csv_name
            )
            log.error(
                "       This product will NOT appear in the dashboard analytics."
            )
            log.error(
                "       Fix: correct the name in the CSV or add SKU to the catalog."
            )
            all_ok = False

    return all_ok  # False = at least one product has no catalog match


# ══════════════════════════════════════════════════════════════════════════════
# DUPLICATE GUARD
# ══════════════════════════════════════════════════════════════════════════════

def _already_ingested(source_id: str) -> bool:
    ds_path = PARQUET_DIR / "daily_sales.parquet"
    if not ds_path.exists():
        return False
    try:
        import polars as pl
        ds = pl.read_parquet(ds_path)
        if "Source" in ds.columns:
            return source_id in ds["Source"].unique().to_list()
    except Exception:
        pass
    return False


# ══════════════════════════════════════════════════════════════════════════════
# SUBPROCESS HELPER
# ══════════════════════════════════════════════════════════════════════════════

def _run(cmd: list[str], cwd: Path, dry_run: bool, label: str) -> bool:
    if dry_run:
        log.info("  [DRY-RUN] %s: %s", label, " ".join(str(c) for c in cmd))
        return True
    try:
        result = subprocess.run(
            cmd, cwd=cwd, capture_output=True, text=True, timeout=120
        )
        if result.returncode == 0:
            log.info("  ✅ %s", label)
            for line in result.stdout.strip().splitlines():
                log.info("     %s", line)
            return True
        else:
            log.error("  ❌ %s failed (exit %d)", label, result.returncode)
            if result.stderr.strip():
                log.error("     %s", result.stderr.strip()[:500])
            return False
    except subprocess.TimeoutExpired:
        log.error("  ❌ %s timed out", label)
        return False
    except Exception as exc:
        log.error("  ❌ %s error: %s", label, exc)
        return False


# ══════════════════════════════════════════════════════════════════════════════
# ARCHIVE
# ══════════════════════════════════════════════════════════════════════════════

def _archive(path: Path, date_start: str, date_end: str, dry_run: bool) -> Path | None:
    canonical = f"item_sales_summary_{date_start}_{date_end}.csv"
    dest = ARCHIVE / canonical
    if dest.exists():
        ts = datetime.datetime.now().strftime("%H%M%S")
        dest = ARCHIVE / f"item_sales_summary_{date_start}_{date_end}_{ts}.csv"
    if dry_run:
        log.info("  [DRY-RUN] archive → data/raw/%s", dest.name)
        return dest
    try:
        shutil.move(str(path), str(dest))
        log.info("  📁 Archived → data/raw/%s", dest.name)
        return dest
    except Exception as exc:
        log.error("  ❌ Archive failed: %s", exc)
        return None


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL MASTER SYNC
# ══════════════════════════════════════════════════════════════════════════════

def _sync_excel_master(source_id: str, date_start: str, date_end: str, dry_run: bool) -> bool:
    if not EXCEL_MASTER.exists():
        log.warning("  ⚠️  Excel Master not found — skipping Excel sync")
        return False
    try:
        import openpyxl
        import polars as pl
    except ImportError as exc:
        log.error("  ❌ Excel sync missing dependency: %s", exc)
        return False

    ds_path = PARQUET_DIR / "daily_sales.parquet"
    cf_path = PARQUET_DIR / "cashflow.parquet"
    if not ds_path.exists():
        log.warning("  ⚠️  daily_sales.parquet not found — skipping Excel sync")
        return False

    ds = pl.read_parquet(ds_path)
    new_sales = ds.filter(pl.col("Source") == source_id) if "Source" in ds.columns else pl.DataFrame()

    if new_sales.is_empty():
        log.info("  ℹ️  No parquet rows for source '%s' — skipping Excel sync", source_id)
        return True

    if dry_run:
        log.info("  [DRY-RUN] Excel sync: %d DailySales rows + Cashflow", new_sales.shape[0])
        return True

    EXCEL_BACKUP.mkdir(parents=True, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = EXCEL_BACKUP / f"FuloFilo_Master_before_{source_id}_{ts}.xlsx"

    try:
        with locked_workbook(EXCEL_MASTER, owner="sales_watcher"):
            shutil.copy2(str(EXCEL_MASTER), str(backup))
            log.info("  💾 Excel backup → backups/%s", backup.name)

            wb = openpyxl.load_workbook(str(EXCEL_MASTER))

            # ── DailySales ────────────────────────────────────────────────────
            ws_sales = wb["DailySales"]
            rows_keep = []
            for i, row in enumerate(ws_sales.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if row and str(row[-1]) != source_id:
                    rows_keep.append(row)
            ws_sales.delete_rows(2, ws_sales.max_row)
            for row in rows_keep:
                ws_sales.append(list(row))

            sku_map: dict[str, str] = {}
            prod_path = PARQUET_DIR / "products.parquet"
            if prod_path.exists():
                prods = pl.read_parquet(prod_path)
                if "full_name" in prods.columns and "sku" in prods.columns:
                    sku_map = dict(
                        zip(prods["full_name"].to_list(), prods["sku"].to_list())
                    )

            appended = 0
            for row in new_sales.iter_rows(named=True):
                sku = sku_map.get(row["Product"], "")
                ws_sales.append([
                    row["Date"],
                    str(sku),
                    row["Product"],
                    round(float(row["Quantity"]), 3),
                    round(float(row["Unit_Price"]), 2),
                    round(float(row["Total"]), 2),
                    row["Payment_Method"],
                    row["Source"],
                ])
                appended += 1
            log.info("  📊 DailySales: +%d rows", appended)

            # ── Cashflow ──────────────────────────────────────────────────────
            if cf_path.exists():
                cf = pl.read_parquet(cf_path)
                ws_cf = wb["Cashflow"]
                prefix = f"Vendas {date_start}"
                cf_keep = []
                for i, row in enumerate(ws_cf.iter_rows(values_only=True)):
                    if i == 0:
                        continue
                    if row and not str(row[3] or "").startswith(prefix):
                        cf_keep.append(row)
                ws_cf.delete_rows(2, ws_cf.max_row)
                for row in cf_keep:
                    ws_cf.append(list(row))
                new_cf = (
                    cf.filter(pl.col("Description").str.starts_with(prefix))
                    if "Description" in cf.columns
                    else pl.DataFrame()
                )
                for row in new_cf.iter_rows(named=True):
                    ws_cf.append([
                        row["Date"],
                        row["Type"],
                        row["Category"],
                        row["Description"],
                        round(float(row["Amount"]), 2),
                        row["Payment_Method"],
                    ])
                log.info("  📊 Cashflow: +%d entries", new_cf.shape[0])

            wb.save(str(EXCEL_MASTER))
            wb.close()
            log.info("  ✅ Excel Master saved")
    except TimeoutError as exc:
        log.error("  ❌ Workbook lock: %s", exc)
        return False

    sync_ok, sync_out = _run_validate_sync_from_excel()
    if not sync_ok:
        log.error(
            "  ❌ sync_excel.sh failed after workbook write — check logs; output:\n%s",
            sync_out[:4000] if sync_out else "(no output)",
        )
        return False
    if sync_out:
        for line in sync_out.strip().splitlines()[:80]:
            log.info("     sync: %s", line)
    log.info("  ✅ Derived artifacts aligned (sync_excel.sh)")
    return True


# ══════════════════════════════════════════════════════════════════════════════
# NOTIFICATIONS + DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

def _notify(title: str, message: str, success: bool = True) -> None:
    """Native macOS notification. Silent fail if unavailable."""
    try:
        sound = "Glass" if success else "Basso"
        script = (
            f'display notification "{message}" '
            f'with title "{title}" '
            f'subtitle "FulôFiló" '
            f'sound name "{sound}"'
        )
        subprocess.run(["osascript", "-e", script], capture_output=True, timeout=5)
    except Exception:
        pass


def _open_dashboard(dry_run: bool) -> None:
    """Open the Streamlit dashboard in the default browser."""
    if dry_run:
        log.info("  [DRY-RUN] open %s", DASHBOARD_URL)
        return
    try:
        subprocess.Popen(["open", DASHBOARD_URL])
        log.info("  🌐 Dashboard opening in browser...")
    except Exception as exc:
        log.warning("  ⚠️  Could not open browser: %s", exc)


# ══════════════════════════════════════════════════════════════════════════════
# GIT
# ══════════════════════════════════════════════════════════════════════════════

def _git_commit_push(archived: Path, date_start: str, date_end: str, dry_run: bool) -> bool:
    msg = f"auto: ingest sales {date_start}"
    if date_start != date_end:
        msg += f" → {date_end}"

    files_to_stage = [
        "data/parquet/",
        "data/excel/FuloFilo_Master.xlsx",
        str(archived.relative_to(ROOT)),
    ]
    ok = _run(["git", "add"] + files_to_stage, cwd=ROOT, dry_run=dry_run, label="git add")
    if not ok:
        return False

    commit_ok = _run(["git", "commit", "-m", msg], cwd=ROOT, dry_run=dry_run, label="git commit")
    if not commit_ok:
        log.info("  ℹ️  Nothing to commit")
        return True

    return _run(["git", "push", "origin", "main"], cwd=ROOT, dry_run=dry_run, label="git push")


# ══════════════════════════════════════════════════════════════════════════════
# CORE PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def process_file(path: Path, dry_run: bool) -> bool:
    cid_token = _correlation_id.set(uuid.uuid4().hex[:12])
    try:
        return _process_file_inner(path, dry_run)
    finally:
        _correlation_id.reset(cid_token)


def _process_file_inner(path: Path, dry_run: bool) -> bool:
    log.info("━" * 60)
    log.info("📥 Processing: %s", path.name)

    # 1 — Validate filename
    m = FILENAME_RE.match(path.name)
    if not m:
        log.warning("  ⏭  Skipping — filename does not match pattern")
        return False
    date_start, date_end = m.group(1), m.group(2)
    source_id = path.stem

    # 2 — Validate internal columns
    if not _has_required_cols(path):
        log.error("  ❌ Invalid columns — file left for manual review")
        _notify("❌ FulôFiló — Ingest Failed", f"Bad columns: {path.name}", success=False)
        return False

    # 3 — Read CSV summary (names + totals) — used in steps 4 and 9
    log.info("  📅 Period: %s → %s", date_start, date_end)
    csv_names, total_units, total_revenue = _read_csv_summary(path)
    log.info("  📦 %d products | %.0f units | R$ %.2f", len(csv_names), total_units, total_revenue)

    # 4 — Fuzzy product name validation
    log.info("  🔍 Validating product names against catalog...")
    known_names = _load_known_product_names()
    names_ok = _fuzzy_validate_products(csv_names, known_names)
    if not names_ok:
        log.warning("  ⚠️  Some products have no catalog match — they will be ingested")
        log.warning("       but won't appear in ABC/revenue analytics until catalog is updated.")
        # Not blocking — user may add new products intentionally

    # 5 — Duplicate guard
    if _already_ingested(source_id):
        log.warning("  ⚠️  '%s' already in parquet — re-ingesting (replaces existing)", source_id)

    # 6 — Run ingest (parquets + analytics rebuild)
    venv_python = ROOT / ".venv" / "bin" / "python"
    python = str(venv_python) if venv_python.exists() else sys.executable
    ingest_args = [python, str(INGEST), str(path)]
    if dry_run:
        ingest_args.append("--dry-run")

    ok = _run(ingest_args, cwd=ROOT, dry_run=False, label="etl/ingest.py")
    if not ok:
        log.error("  ❌ Ingest failed — file left in incoming/ for review")
        _notify("❌ FulôFiló — Ingest Failed", f"Error: {path.name}", success=False)
        return False

    # 7 — Sync Excel Master
    log.info("  📋 Syncing Excel Master...")
    if not _sync_excel_master(source_id, date_start, date_end, dry_run):
        log.error("  ❌ Excel write or post-sync failed — file left in incoming/")
        _notify(
            "❌ FulôFiló — Excel / sync failed",
            f"Check logs: {path.name}",
            success=False,
        )
        return False

    # 8 — Archive
    archived = _archive(path, date_start, date_end, dry_run)
    if not archived:
        return False

    # 9 — Git commit + push
    pushed = _git_commit_push(archived, date_start, date_end, dry_run)

    # 10 — Notify + open dashboard
    period_label = date_start if date_start == date_end else f"{date_start} → {date_end}"
    if pushed:
        log.info("  🚀 Done — Streamlit Cloud redeploys in ~90 seconds")
        _notify(
            "✅ FulôFiló — Dashboard updated!",
            f"{period_label} | {total_units:.0f} un | R$ {total_revenue:,.2f} | Live in ~90s"
        )
        _open_dashboard(dry_run)
    else:
        _notify("⚠️ FulôFiló — Push failed", "Check logs/saleswatch.log", success=False)

    return pushed


# ══════════════════════════════════════════════════════════════════════════════
# SCAN LOOP
# ══════════════════════════════════════════════════════════════════════════════

def scan_and_process(dry_run: bool) -> int:
    if not INCOMING.exists():
        INCOMING.mkdir(parents=True)
        log.info("Created incoming folder: %s", INCOMING)

    candidates = sorted(INCOMING.glob("item-sales-summary-*.csv"))
    if not candidates:
        log.debug("No matching files in %s", INCOMING)
        return 0

    log.info("Found %d file(s) to process", len(candidates))
    ok_count = 0
    for f in candidates:
        if process_file(f, dry_run):
            ok_count += 1

    log.info("━" * 60)
    log.info("✅ Processed %d / %d file(s)", ok_count, len(candidates))
    return ok_count


def loop_mode(interval: int = 30, dry_run: bool = False) -> None:
    """Persistent daemon mode — polls every N seconds."""
    import time
    log.info("=" * 60)
    log.info("FulôFiló Sales Watcher — DAEMON (poll every %ds)", interval)
    log.info("Watch folder: %s", INCOMING)
    log.info("=" * 60)
    while True:
        try:
            scan_and_process(dry_run=dry_run)
        except Exception as exc:
            log.error("Unhandled error in scan loop: %s", exc)
        time.sleep(interval)


def main() -> None:
    parser = argparse.ArgumentParser(description="FulôFiló — Sales Drop Watcher v3")
    parser.add_argument("--dry-run", action="store_true",
                        help="Simulate — no files written, no git push, no browser")
    parser.add_argument("--daemon", action="store_true",
                        help="Poll incoming/ forever (LaunchAgent / long-running host)")
    parser.add_argument("--interval", type=int, default=30,
                        help="Seconds between scans in --daemon mode (default: 30)")
    args = parser.parse_args()

    log.info("=" * 60)
    log.info("FulôFiló Sales Watcher v3  %s%s",
             datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
             "  [DRY-RUN]" if args.dry_run else "")
    log.info("Watch folder: %s", INCOMING)
    log.info("=" * 60)

    if args.daemon:
        loop_mode(args.interval, dry_run=args.dry_run)
    else:
        scan_and_process(args.dry_run)


if __name__ == "__main__":
    main()
