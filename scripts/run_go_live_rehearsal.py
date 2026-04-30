#!/usr/bin/env python3
"""
FulôFiló — Phase 5 go-live rehearsal runner
===========================================
Creates a sanitized non-production workbook fixture and exercises the
canonical Excel-first workflow end to end without mutating the real
source-of-truth workbook.
"""

from __future__ import annotations

import json
import subprocess
import sys
import time
import urllib.error
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
PYTHON = ROOT / ".venv" / "bin" / "python3"
STREAMLIT = ROOT / ".venv" / "bin" / "streamlit"
REHEARSAL_DIR = ROOT / "data" / "excel" / "rehearsal"
WORKBOOK_PATH = REHEARSAL_DIR / "FuloFilo_Rehearsal_NON_PRODUCTION.xlsx"
RESULT_PATH = REHEARSAL_DIR / "go_live_rehearsal_report.json"
STATUS_PATH = ROOT / "data" / "excel" / "source_sync_status.json"
REPORT_PATH = ROOT / "excel" / f"FuloFilo_Rehearsal_Report_{date.today().isoformat()}.xlsx"
PORT = 8510


def _run(cmd: list[str], check: bool = True) -> subprocess.CompletedProcess[str]:
    result = subprocess.run(cmd, cwd=str(ROOT), capture_output=True, text=True)
    if check and result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or result.stdout.strip() or f"Command failed: {' '.join(cmd)}")
    return result


def build_rehearsal_workbook(output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Catalog")
    ws.append(["sku", "full_name", "category", "unit_cost", "suggested_price", "min_stock", "reorder_qty"])
    catalog_rows = [
        ["10001", "Bolsa Praia Listrada", "Acessórios", 24.0, 59.0, 6, 18],
        ["10002", "Canga Sol do Nordeste", "Cangas", 38.0, 89.0, 4, 12],
        ["10003", "Caneca Tropical Azul", "Souvenirs", 11.5, 29.9, 8, 24],
        ["10004", "Chaveiro Mar Bordado", "Souvenirs", 5.0, 14.9, 12, 36],
        ["10005", "Necessaire Coqueiro", "Acessórios", 17.5, 39.9, 5, 15],
        ["10006", "Camiseta Nordeste Arte", "Vestuário", 28.0, 69.9, 3, 9],
    ]
    for row in catalog_rows:
        ws.append(row)

    ws = wb.create_sheet("Inventory")
    ws.append(["sku", "product", "category", "current_stock", "min_stock", "reorder_qty", "supplier", "lead_time_days", "notes"])
    inventory_rows = [
        ["10001", "Bolsa Praia Listrada", "Acessórios", 18, 6, 18, "Fornecedor Praia", 7, "Rehearsal data"],
        ["10002", "Canga Sol do Nordeste", "Cangas", 11, 4, 12, "Fornecedor Cangas", 10, "Rehearsal data"],
        ["10003", "Caneca Tropical Azul", "Souvenirs", 32, 8, 24, "Fornecedor Canecas", 12, "Rehearsal data"],
        ["10004", "Chaveiro Mar Bordado", "Souvenirs", 45, 12, 36, "Fornecedor Chaveiros", 5, "Rehearsal data"],
        ["10005", "Necessaire Coqueiro", "Acessórios", 16, 5, 15, "Fornecedor Bolsas", 8, "Rehearsal data"],
        ["10006", "Camiseta Nordeste Arte", "Vestuário", 9, 3, 9, "Fornecedor Vestuário", 14, "Rehearsal data"],
    ]
    for row in inventory_rows:
        ws.append(row)

    ws = wb.create_sheet("DailySales")
    ws.append(["Date", "sku", "Product", "Quantity", "Unit_Price", "Total", "Payment_Method", "Source"])
    sales_rows = [
        ["2026-04-20", "10001", "Bolsa Praia Listrada", 3, 59.0, 177.0, "Pix", "rehearsal"],
        ["2026-04-20", "10003", "Caneca Tropical Azul", 4, 29.9, 119.6, "Débito", "rehearsal"],
        ["2026-04-21", "10002", "Canga Sol do Nordeste", 2, 89.0, 178.0, "Crédito", "rehearsal"],
        ["2026-04-22", "10005", "Necessaire Coqueiro", 2, 39.9, 79.8, "Pix", "rehearsal"],
        ["2026-04-23", "10006", "Camiseta Nordeste Arte", 1, 69.9, 69.9, "Dinheiro", "rehearsal"],
    ]
    for row in sales_rows:
        ws.append(row)

    ws = wb.create_sheet("Cashflow")
    ws.append(["Date", "Type", "Category", "Description", "Amount", "Payment_Method"])
    cashflow_rows = [
        ["2026-04-20", "Receita", "Vendas", "Resumo vendas 20/04", 296.6, "Pix"],
        ["2026-04-21", "Receita", "Vendas", "Resumo vendas 21/04", 178.0, "Cartão"],
        ["2026-04-22", "Despesa", "Fornecedores", "Compra de estoque", 180.0, "Boleto"],
        ["2026-04-23", "Despesa", "Operação", "Frete interno", 35.0, "Pix"],
    ]
    for row in cashflow_rows:
        ws.append(row)

    ws = wb.create_sheet("CategoryOverrides")
    ws.append(["sku", "category", "subcategory", "confidence"])
    ws.append(["10006", "Vestuário", "Camisetas", "manual"])

    ws = wb.create_sheet("Meta")
    ws.append(["key", "value"])
    ws.append(["schema_version", "1"])
    ws.append(["workbook", "FuloFilo_Rehearsal_NON_PRODUCTION"])
    ws.append(["environment", "rehearsal"])
    ws.append(["data_label", "SANITIZED_NON_PRODUCTION"])

    wb.save(output_path)
    return output_path


def sync_rehearsal_workbook(workbook_path: Path = WORKBOOK_PATH) -> tuple[bool, str]:
    result = _run([str(PYTHON), "scripts/sync_excel.py", "--excel", str(workbook_path)], check=False)
    output = "\n".join(part for part in [result.stdout.strip(), result.stderr.strip()] if part).strip()
    return result.returncode == 0, output


def read_status() -> dict:
    return json.loads(STATUS_PATH.read_text(encoding="utf-8"))


def get_dashboard_snapshot() -> dict:
    from app.db import get_abc_analysis, get_conn, get_margin_matrix, get_summary_kpis

    conn = get_conn()
    try:
        summary = get_summary_kpis(conn, "ALL")
        abc = get_abc_analysis(conn, "ALL")
        margin = get_margin_matrix(conn, "ALL")
    finally:
        conn.close()
    return {
        "summary": summary,
        "abc_rows": abc.height,
        "margin_rows": margin.height,
    }


def start_dashboard_smoke() -> dict:
    proc = subprocess.Popen(
        [
            str(STREAMLIT),
            "run",
            "app/app.py",
            "--server.port",
            str(PORT),
            "--server.address",
            "127.0.0.1",
            "--server.headless",
            "true",
        ],
        cwd=str(ROOT),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )

    health_url = f"http://127.0.0.1:{PORT}/_stcore/health"
    started = False
    error = ""
    try:
        for _ in range(30):
            if proc.poll() is not None:
                stderr = proc.stderr.read() if proc.stderr else ""
                stdout = proc.stdout.read() if proc.stdout else ""
                error = (stderr or stdout).strip()
                break
            try:
                with urllib.request.urlopen(health_url, timeout=1) as response:
                    if response.status == 200:
                        started = True
                        break
            except (urllib.error.URLError, TimeoutError):
                time.sleep(1)
        return {"started": started, "url": f"http://127.0.0.1:{PORT}", "error": error}
    finally:
        proc.terminate()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()


def restore_canonical_outputs() -> tuple[bool, str]:
    result = _run(["bash", "scripts/sync_excel.sh"], check=False)
    output = "\n".join(part for part in [result.stdout.strip(), result.stderr.strip()] if part).strip()
    return result.returncode == 0, output


def main() -> int:
    from app.utils.category_ops import upsert_category_override
    from app.utils.inventory_ops import adjust_stock
    from app.utils.sales_ops import append_sale_to_excel
    from excel.build_report import build_report

    results: dict[str, object] = {
        "workbook": str(WORKBOOK_PATH),
        "report": str(REPORT_PATH),
        "steps": {},
        "warnings": [],
        "issues": [],
    }

    workbook = build_rehearsal_workbook(WORKBOOK_PATH)
    results["workbook_created"] = str(workbook)

    try:
        ok, sync_output = sync_rehearsal_workbook(workbook)
        results["steps"]["initial_sync"] = {"ok": ok, "output": sync_output}
        if not ok:
            raise RuntimeError(sync_output or "Initial rehearsal sync failed.")

        status = read_status()
        results["initial_status"] = status
        if not status.get("healthy_production_data"):
            raise RuntimeError(f"Rehearsal workbook did not become healthy: {status}")

        snapshot = get_dashboard_snapshot()
        results["dashboard_snapshot_before"] = snapshot
        receita, quantidade, lucro, ticket = snapshot["summary"]
        if not receita or not quantidade:
            raise RuntimeError(f"Dashboard KPIs stayed empty after rehearsal sync: {snapshot}")

        dashboard_start = start_dashboard_smoke()
        results["steps"]["dashboard_launch"] = dashboard_start
        if not dashboard_start["started"]:
            raise RuntimeError(dashboard_start["error"] or "Dashboard did not reach health endpoint.")

        def sync_runner() -> tuple[bool, str]:
            return sync_rehearsal_workbook(workbook)

        sale_result = append_sale_to_excel(
            sale_date=date(2026, 4, 27),
            sku="10004",
            product="Chaveiro Mar Bordado",
            quantity=3,
            unit_price=14.9,
            payment_method="Pix",
            source="rehearsal-app",
            workbook_path=workbook,
            sync_runner=sync_runner,
        )
        results["sale_writeback"] = sale_result.__dict__

        inventory_result = adjust_stock(
            sku="10002",
            new_qty=18,
            workbook_path=workbook,
            sync_runner=sync_runner,
        )
        results["inventory_adjustment"] = inventory_result.__dict__

        category_result = upsert_category_override(
            sku="10003",
            category="Presentes",
            subcategory="Canecas",
            confidence="manual",
            workbook_path=workbook,
            sync_runner=sync_runner,
        )
        results["category_override"] = category_result.__dict__

        final_status = read_status()
        results["final_status"] = final_status
        if not final_status.get("healthy_production_data"):
            raise RuntimeError(f"Rehearsal lost healthy status after write-backs: {final_status}")

        final_snapshot = get_dashboard_snapshot()
        results["dashboard_snapshot_after"] = final_snapshot

        report_path = build_report(output_path=REPORT_PATH)
        results["steps"]["report_export"] = {"ok": report_path.exists(), "path": str(report_path)}
        if not report_path.exists():
            raise RuntimeError("Rehearsal report was not created.")

        results["warnings"] = final_status.get("warnings", [])
        RESULT_PATH.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"[rehearsal] Workbook: {workbook}")
        print(f"[rehearsal] Report: {report_path}")
        print(f"[rehearsal] Results: {RESULT_PATH}")
        return 0
    except Exception as exc:  # noqa: BLE001
        results["issues"].append(str(exc))
        RESULT_PATH.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"[rehearsal] FAILED: {exc}", file=sys.stderr)
        return 1
    finally:
        restore_canonical_outputs()


if __name__ == "__main__":
    raise SystemExit(main())
