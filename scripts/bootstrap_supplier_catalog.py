#!/usr/bin/env python3
"""Populate SupplierCatalog + PurchaseOrders sheets from catalog/inventory and suppliers.json."""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.utils.excel_sync import backup_workbook

XLSX = ROOT / "data" / "excel" / "FuloFilo_Master.xlsx"
SUPPLIERS_JSON = ROOT / "data" / "suppliers" / "suppliers.json"

SUPPLIER_CATALOG_HEADERS = [
    "sku", "supplier_id", "supplier_name", "lead_time_days", "moq", "case_pack",
]
PURCHASE_ORDER_HEADERS = [
    "po_id", "supplier_id", "supplier_name", "sku", "product", "qty",
    "unit_cost", "line_total", "status", "created_at", "notes",
]

# Category -> (supplier_id, supplier_name, lead_time_days, moq, case_pack)
CATEGORY_MAP = {
    "Roupas": ("2", "ALGODOEIRO ECO FASHION", 21, 0, 1),
    "Cangas em Elastano": ("4", "ALAN BOLSAS", 45, 6, 6),
    "Cangas em Algodão": ("9", "GIRISH INDIANA", 30, 0, 1),
    "Bolsas": ("4", "ALAN BOLSAS", 45, 0, 1),
    "Outros": ("3", "DISTRIBUIDOR (GENÉRICO)", 14, 0, 1),
    "Sem categoria": ("3", "DISTRIBUIDOR (GENÉRICO)", 14, 0, 1),
}
DEFAULT_SUPPLIER = ("3", "DISTRIBUIDOR (GENÉRICO)", 14, 0, 1)


def _norm_sku(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    try:
        return str(int(float(s))).zfill(5)
    except ValueError:
        return s


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Workbook not found: {XLSX}")

    backup_workbook(workbook_path=XLSX)
    wb = openpyxl.load_workbook(XLSX)

    if "SupplierCatalog" not in wb.sheetnames:
        wb.create_sheet("SupplierCatalog")
    sc_ws = wb["SupplierCatalog"]
    sc_ws.delete_rows(1, sc_ws.max_row)
    sc_ws.append(SUPPLIER_CATALOG_HEADERS)

    if "PurchaseOrders" not in wb.sheetnames:
        wb.create_sheet("PurchaseOrders")
    po_ws = wb["PurchaseOrders"]
    if po_ws.max_row == 0 or str(po_ws.cell(1, 1).value or "") != "po_id":
        po_ws.delete_rows(1, po_ws.max_row)
        po_ws.append(PURCHASE_ORDER_HEADERS)

    catalog = {row[0]: row for row in wb["Catalog"].iter_rows(min_row=2, values_only=True) if row[0]}
    inv_header = [c.value for c in wb["Inventory"][1]]
    inv_rows = list(wb["Inventory"].iter_rows(min_row=2, values_only=True))

    supplier_col = inv_header.index("supplier") if "supplier" in inv_header else None
    lead_col = inv_header.index("lead_time_days") if "lead_time_days" in inv_header else None

    rows_written = 0
    for inv_row in inv_rows:
        if not inv_row or not inv_row[0]:
            continue
        sku = _norm_sku(inv_row[0])
        product = inv_row[1]
        category = inv_row[2] if len(inv_row) > 2 else ""

        existing_supplier = ""
        existing_lead = 7
        if supplier_col is not None and inv_row[supplier_col]:
            existing_supplier = str(inv_row[supplier_col]).strip()
        if lead_col is not None and inv_row[lead_col]:
            try:
                existing_lead = int(inv_row[lead_col])
            except (TypeError, ValueError):
                existing_lead = 7

        sid, sname, lead, moq, pack = CATEGORY_MAP.get(str(category or ""), DEFAULT_SUPPLIER)
        if existing_lead and existing_lead != 7:
            lead = existing_lead

        sc_ws.append([sku, sid, sname, lead, moq, pack])
        rows_written += 1

        if supplier_col is not None and not existing_supplier:
            row_idx = inv_rows.index(inv_row) + 2
            wb["Inventory"].cell(row=row_idx, column=supplier_col + 1, value=sname)

    wb.save(XLSX)
    print(f"✅ SupplierCatalog: {rows_written} SKU mappings written to {XLSX}")
    if SUPPLIERS_JSON.exists():
        data = json.loads(SUPPLIERS_JSON.read_text(encoding="utf-8"))
        print(f"   Linked to {len(data.get('suppliers', []))} suppliers in suppliers.json")


if __name__ == "__main__":
    main()
