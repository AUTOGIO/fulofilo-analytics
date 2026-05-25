"""Tests for voice memo → Canga transcript parsing."""

from __future__ import annotations

import csv
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from etl.parse_voice_memo_transcripts import (  # noqa: E402
    CatalogRow,
    load_catalog,
    parse_lines,
)


@pytest.fixture
def tiny_catalog() -> list[CatalogRow]:
    return [
        CatalogRow(
            sku=401,
            tipo="ELASTANO",
            estampa_canonica="Olho Grego",
            aliases=["olho grego", "olho_grego", "Olho Grego"],
        ),
        CatalogRow(
            sku=429,
            tipo="ALGODAO",
            estampa_canonica="Cordel",
            aliases=["cordel", "Cordel"],
        ),
    ]


def test_parse_line_elastano_quantity(tiny_catalog: list[CatalogRow]) -> None:
    rows = parse_lines(tiny_catalog, "elastano olho grego 3\n")
    assert len(rows) == 1
    assert rows[0]["tipo_canga"] == "ELASTANO"
    assert rows[0]["estampa"] == "Olho Grego"
    assert rows[0]["quantidade"] == 3
    assert rows[0]["sku"] == 401
    assert rows[0]["confidence"] >= 0.85
    assert rows[0]["needs_review"] is False


def test_parse_line_alg_word_number(tiny_catalog: list[CatalogRow]) -> None:
    rows = parse_lines(tiny_catalog, "algodao cordel cinco")
    assert rows[0]["tipo_canga"] == "ALGODAO"
    assert rows[0]["quantidade"] == 5
    assert rows[0]["sku"] == 429


def test_uma_duzia_quantity(tiny_catalog: list[CatalogRow]) -> None:
    rows = parse_lines(tiny_catalog, "elastano olho grego uma duzia")
    assert rows[0]["quantidade"] == 12
    assert rows[0]["sku"] == 401


def test_memo_title_default_tipo(tiny_catalog: list[CatalogRow]) -> None:
    text = """--- memo_begin 2026-04-30T10:00:00 Pedido elastano ---
olho grego 2
--- memo_end ---
"""
    rows = parse_lines(tiny_catalog, text)
    assert len(rows) == 1
    assert rows[0]["tipo_canga"] == "ELASTANO"
    assert rows[0]["quantidade"] == 2
    assert rows[0]["sku"] == 401


def test_exported_catalog_csv_readable() -> None:
    path = ROOT / "data" / "raw" / "voice_memo_transcripts" / "canga_catalog_400_438.csv"
    if not path.exists():
        pytest.skip("catalog not generated; run export_canga_voice_memo_catalog.py")
    rows = load_catalog(path)
    assert len(rows) >= 1
    skus = {r.sku for r in rows}
    assert 401 in skus or 400 in skus


def test_export_script_creates_csv(tmp_path: Path) -> None:
    import openpyxl

    from etl.export_canga_voice_memo_catalog import main as export_main

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalog"
    ws.append(["sku", "full_name", "category", "unit_cost", "suggested_price", "min_stock", "reorder_qty"])
    ws.append([401, "Canga Elastano — Olho Grego", "Cangas em Elastano", 1, 2, 3, 4])
    xlsx = tmp_path / "mini.xlsx"
    wb.save(xlsx)
    out = tmp_path / "out.csv"

    import sys as _sys

    old_argv = _sys.argv
    try:
        _sys.argv = ["export_canga_voice_memo_catalog.py", "--master", str(xlsx), "-o", str(out)]
        export_main()
    finally:
        _sys.argv = old_argv

    with out.open(encoding="utf-8") as f:
        r = list(csv.DictReader(f))
    assert len(r) == 1
    assert r[0]["sku"] == "401"
    assert r[0]["tipo"] == "ELASTANO"
    assert "Olho Grego" in r[0]["estampa_canonica"]


def test_write_parsed_xlsx_sheets_and_summary(tmp_path: Path) -> None:
    import openpyxl

    from etl.parse_voice_memo_transcripts import PARSED_XLSX_COLUMNS, write_parsed_xlsx

    rows = [
        {
            "tipo_canga": "ELASTANO",
            "estampa": "Olho Grego",
            "quantidade": 1,
            "sku": 401,
            "confidence": 1.0,
            "needs_review": False,
            "raw_line": "elastano olho grego 1",
            "source_file": "day.txt",
        },
        {
            "tipo_canga": "ALGODAO",
            "estampa": "Cordel",
            "quantidade": 2,
            "sku": 429,
            "confidence": 0.8,
            "needs_review": True,
            "raw_line": "algodao cordel 2",
            "source_file": "day.txt",
        },
    ]
    out = tmp_path / "parsed.xlsx"
    write_parsed_xlsx(rows, out)
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    assert set(wb.sheetnames) >= {"Parsed", "Summary"}
    ws = wb["Parsed"]
    assert [c.value for c in ws[1]] == PARSED_XLSX_COLUMNS
    assert ws.max_row == 3
    summary = list(wb["Summary"].iter_rows(values_only=True))
    wb.close()
    assert ("parsed_row_count", 2) in summary
    assert ("needs_review_true_count", 1) in summary


def _minimal_master_for_merge(path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    wc = wb.active
    assert wc is not None
    wc.title = "Catalog"
    wc.append(["sku", "full_name", "category", "unit_cost", "suggested_price", "min_stock", "reorder_qty"])
    wc.append([401, "Canga Elastano — Olho Grego", "Cangas em Elastano", 1.0, 2.0, 10, 20])
    wc.append([402, "Canga Elastano — Other", "Cangas em Elastano", 1.0, 2.0, 10, 20])

    wi = wb.create_sheet("Inventory")
    wi.append(
        [
            "sku",
            "product",
            "category",
            "current_stock",
            "min_stock",
            "reorder_qty",
            "supplier",
            "lead_time_days",
            "notes",
        ]
    )
    wi.append([401, "Canga Elastano — Olho Grego", "Cangas em Elastano", 100, 10, 20, None, 7, None])
    wb.save(path)


def test_merge_plan_increment_and_append(tmp_path: Path) -> None:
    import openpyxl

    from etl.merge_voice_memo_parsed_to_inventory import aggregate_qty_by_sku, plan_inventory_merge

    master = tmp_path / "FuloFilo_Master.xlsx"
    _minimal_master_for_merge(master)
    parsed = [
        {"sku": "401", "quantidade": "7"},
        {"sku": 402, "quantidade": 3},
    ]
    deltas = aggregate_qty_by_sku(parsed)
    assert deltas == {401: 7, 402: 3}
    changes = plan_inventory_merge(master, deltas)
    by_action = {c["sku"]: c for c in changes}
    assert by_action[401]["action"] == "increment"
    assert by_action[401]["old_current_stock"] == 100
    assert by_action[401]["new_current_stock"] == 107
    assert by_action[402]["action"] == "append"
    assert by_action[402]["current_stock"] == 3


def test_merge_apply_and_backup(tmp_path: Path) -> None:
    import openpyxl

    from etl.merge_voice_memo_parsed_to_inventory import (
        aggregate_qty_by_sku,
        apply_inventory_merge,
        backup_master,
        plan_inventory_merge,
    )

    master = tmp_path / "FuloFilo_Master.xlsx"
    _minimal_master_for_merge(master)
    deltas = aggregate_qty_by_sku([{"sku": 401, "quantidade": 5}, {"sku": 402, "quantidade": 1}])
    changes = [c for c in plan_inventory_merge(master, deltas) if c["action"] in ("increment", "append")]
    bk_dir = tmp_path / "backups"
    backup_master(master, bk_dir)
    apply_inventory_merge(master, changes)
    wb = openpyxl.load_workbook(master, data_only=True)
    inv = wb["Inventory"]
    rows = {inv.cell(row=r, column=1).value: inv.cell(row=r, column=4).value for r in range(2, inv.max_row + 1)}
    wb.close()
    assert rows[401] == 105
    assert rows[402] == 1
    assert list(bk_dir.glob("FuloFilo_Master_*_before_voice_memo_inventory_merge.xlsx"))


def test_merge_dry_run_skips_backup(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    import sys

    from etl import merge_voice_memo_parsed_to_inventory as m

    master = tmp_path / "FuloFilo_Master.xlsx"
    _minimal_master_for_merge(master)
    csv_path = tmp_path / "p.csv"
    csv_path.write_text(
        "tipo_canga,estampa,quantidade,sku,confidence,needs_review,source_file,raw_line\n"
        "ELASTANO,Olho Grego,2,401,1.0,False,t.txt,line\n",
        encoding="utf-8",
    )
    old = sys.argv
    try:
        sys.argv = [
            "merge_voice_memo_parsed_to_inventory.py",
            "--parsed-csv",
            str(csv_path),
            "--master",
            str(master),
            "--dry-run",
        ]
        m.main()
    finally:
        sys.argv = old
    captured = capsys.readouterr().out
    assert "INCREMENT" in captured
    assert "[dry-run]" in captured
    assert not list(tmp_path.glob("**/FuloFilo_Master_*_before_voice_memo_inventory_merge.xlsx"))


def test_load_parsed_xlsx_roundtrip(tmp_path: Path) -> None:
    from etl.merge_voice_memo_parsed_to_inventory import load_parsed_xlsx
    from etl.parse_voice_memo_transcripts import write_parsed_xlsx

    rows = [
        {
            "tipo_canga": "ELASTANO",
            "estampa": "Olho Grego",
            "quantidade": 4,
            "sku": 401,
            "confidence": 0.99,
            "needs_review": False,
            "raw_line": "x",
            "source_file": "s.txt",
        }
    ]
    x = tmp_path / "r.xlsx"
    write_parsed_xlsx(rows, x)
    loaded = load_parsed_xlsx(x)
    assert len(loaded) == 1
    assert int(loaded[0]["sku"]) == 401
    assert loaded[0]["quantidade"] == 4


def test_merge_aggregate_skip_needs_review_csv_strings(tmp_path: Path) -> None:
    from etl.merge_voice_memo_parsed_to_inventory import aggregate_qty_by_sku, plan_inventory_merge

    master = tmp_path / "FuloFilo_Master.xlsx"
    _minimal_master_for_merge(master)
    parsed = [
        {"sku": "401", "quantidade": "10", "needs_review": "TRUE"},
        {"sku": "401", "quantidade": "2", "needs_review": "FALSE"},
        {"sku": "401", "quantidade": "3", "needs_review": "false"},
        {"sku": "401", "quantidade": "1", "needs_review": "sim"},
    ]
    assert aggregate_qty_by_sku(parsed) == {401: 16}
    assert aggregate_qty_by_sku(parsed, skip_needs_review=True) == {401: 5}
    changes = plan_inventory_merge(master, aggregate_qty_by_sku(parsed, skip_needs_review=True))
    by_sku = {c["sku"]: c for c in changes}
    assert by_sku[401]["action"] == "increment"
    assert by_sku[401]["delta"] == 5
    assert by_sku[401]["new_current_stock"] == 105


def test_merge_skip_needs_review_main_dry_run(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    import sys

    from etl import merge_voice_memo_parsed_to_inventory as m

    master = tmp_path / "FuloFilo_Master.xlsx"
    _minimal_master_for_merge(master)
    csv_path = tmp_path / "p.csv"
    csv_path.write_text(
        "tipo_canga,estampa,quantidade,sku,confidence,needs_review,source_file,raw_line\n"
        "ELASTANO,Olho Grego,9,401,1.0,True,t.txt,line1\n"
        "ELASTANO,Olho Grego,2,401,1.0,False,t.txt,line2\n",
        encoding="utf-8",
    )
    old = sys.argv
    try:
        sys.argv = [
            "merge_voice_memo_parsed_to_inventory.py",
            "--parsed-csv",
            str(csv_path),
            "--master",
            str(master),
            "--dry-run",
            "--skip-needs-review",
        ]
        m.main()
    finally:
        sys.argv = old
    out = capsys.readouterr().out
    assert "Rows included in aggregation (--skip-needs-review): 1" in out
    assert "+ 2 -> 102" in out
    assert "+ 9 -> 109" not in out


def test_merge_aggregate_skip_needs_review_xlsx(tmp_path: Path) -> None:
    from etl.merge_voice_memo_parsed_to_inventory import aggregate_qty_by_sku, load_parsed_xlsx
    from etl.parse_voice_memo_transcripts import write_parsed_xlsx

    rows = [
        {
            "tipo_canga": "ELASTANO",
            "estampa": "Olho Grego",
            "quantidade": 8,
            "sku": 401,
            "confidence": 0.5,
            "needs_review": True,
            "raw_line": "x",
            "source_file": "s.txt",
        },
        {
            "tipo_canga": "ELASTANO",
            "estampa": "Olho Grego",
            "quantidade": 4,
            "sku": 401,
            "confidence": 0.99,
            "needs_review": False,
            "raw_line": "y",
            "source_file": "s.txt",
        },
    ]
    x = tmp_path / "mix.xlsx"
    write_parsed_xlsx(rows, x)
    loaded = load_parsed_xlsx(x)
    assert aggregate_qty_by_sku(loaded) == {401: 12}
    assert aggregate_qty_by_sku(loaded, skip_needs_review=True) == {401: 4}
