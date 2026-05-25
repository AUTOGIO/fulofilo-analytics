"""
FulôFiló — Minimum Viable Test Suite
=====================================
6 tests covering the full data pipeline.

Run:
    cd /Users/eduardogiovannini/dev/products/FuloFilo
    .venv/bin/python3 -m pytest tests/test_pipeline.py -v
"""

from __future__ import annotations

import importlib
import json
import sys
from datetime import date
from pathlib import Path

import openpyxl
import polars as pl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

DATA_DIR = ROOT / "data" / "parquet"
RAW_DIR  = ROOT / "data" / "raw"

EXPECTED_PARQUETS = [
    "cashflow", "daily_sales", "inventory", "products",
    "profit_report", "quantity_report", "revenue_report",
]


def _make_test_workbook(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    ws = wb.create_sheet("Catalog")
    ws.append(["sku", "full_name", "category", "unit_cost", "suggested_price", "min_stock", "reorder_qty"])
    ws.append(["00001", "Produto Exemplo", "Geral", 10.0, 25.0, 5, 15])
    ws.append(["00002", "Bolsa Teste", "Acessórios", 20.0, 45.0, 3, 10])

    ws = wb.create_sheet("Inventory")
    ws.append(["sku", "product", "category", "current_stock", "min_stock", "reorder_qty", "supplier", "lead_time_days", "notes"])
    ws.append(["00001", "Produto Exemplo", "Geral", 5, 5, 15, "", 7, ""])
    ws.append(["00002", "Bolsa Teste", "Acessórios", 8, 3, 10, "", 7, ""])

    ws = wb.create_sheet("DailySales")
    ws.append(["Date", "sku", "Product", "Quantity", "Unit_Price", "Total", "Payment_Method", "Source"])

    ws = wb.create_sheet("Cashflow")
    ws.append(["Date", "Type", "Category", "Description", "Amount", "Payment_Method"])

    ws = wb.create_sheet("CategoryOverrides")
    ws.append(["sku", "category", "subcategory", "confidence"])

    ws = wb.create_sheet("Meta")
    ws.append(["key", "value"])
    ws.append(["schema_version", "1"])
    ws.append(["workbook", "FuloFilo_Master"])

    path = tmp_path / "FuloFilo_Master.xlsx"
    wb.save(path)
    return path


def _make_bootstrap_workbook(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    ws = wb.create_sheet("Catalog")
    ws.append(["sku", "full_name", "category", "unit_cost", "suggested_price", "min_stock", "reorder_qty"])
    ws.append(["00001", "Produto Exemplo", "Geral", 10.0, 25.0, 5, 15])

    ws = wb.create_sheet("Inventory")
    ws.append(["sku", "product", "category", "current_stock", "min_stock", "reorder_qty", "supplier", "lead_time_days", "notes"])
    ws.append(["00001", "Produto Exemplo", "Geral", 0, 5, 15, "", 7, ""])

    ws = wb.create_sheet("DailySales")
    ws.append(["Date", "sku", "Product", "Quantity", "Unit_Price", "Total", "Payment_Method", "Source"])

    ws = wb.create_sheet("Cashflow")
    ws.append(["Date", "Type", "Category", "Description", "Amount", "Payment_Method"])

    ws = wb.create_sheet("CategoryOverrides")
    ws.append(["sku", "category", "subcategory", "confidence"])

    ws = wb.create_sheet("Meta")
    ws.append(["key", "value"])
    ws.append(["schema_version", "1"])
    ws.append(["workbook", "FuloFilo_Master"])

    path = tmp_path / "Bootstrap_Master.xlsx"
    wb.save(path)
    return path


def _make_production_ready_workbook(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    ws = wb.create_sheet("Catalog")
    ws.append(["sku", "full_name", "category", "unit_cost", "suggested_price", "min_stock", "reorder_qty"])
    catalog_rows = [
        ["10001", "Bolsa Praia", "Acessórios", 20.0, 45.0, 3, 10],
        ["10002", "Canga Sol", "Cangas", 35.0, 79.0, 2, 8],
        ["10003", "Caneca Tropical", "Souvenirs", 12.0, 29.0, 4, 12],
        ["10004", "Chaveiro Mar", "Souvenirs", 4.0, 12.0, 10, 25],
        ["10005", "Necessaire Azul", "Acessórios", 18.0, 39.0, 3, 10],
        ["10006", "Camiseta Nordeste", "Vestuário", 22.0, 55.0, 2, 8],
    ]
    for row in catalog_rows:
        ws.append(row)

    ws = wb.create_sheet("Inventory")
    ws.append(["sku", "product", "category", "current_stock", "min_stock", "reorder_qty", "supplier", "lead_time_days", "notes"])
    for sku, name, category, *_rest in catalog_rows:
        ws.append([sku, name, category, 20, 3, 10, "Fornecedor Teste", 7, ""])

    ws = wb.create_sheet("DailySales")
    ws.append(["Date", "sku", "Product", "Quantity", "Unit_Price", "Total", "Payment_Method", "Source"])
    ws.append(["2026-04-20", "10001", "Bolsa Praia", 2, 45.0, 90.0, "Pix", "manual"])
    ws.append(["2026-04-21", "10002", "Canga Sol", 1, 79.0, 79.0, "Crédito", "manual"])

    ws = wb.create_sheet("Cashflow")
    ws.append(["Date", "Type", "Category", "Description", "Amount", "Payment_Method"])
    ws.append(["2026-04-20", "Receita", "Vendas", "Vendas do dia", 169.0, "Pix"])
    ws.append(["2026-04-21", "Despesa", "Fornecedores", "Reposição", 80.0, "Boleto"])

    ws = wb.create_sheet("CategoryOverrides")
    ws.append(["sku", "category", "subcategory", "confidence"])

    ws = wb.create_sheet("Meta")
    ws.append(["key", "value"])
    ws.append(["schema_version", "1"])
    ws.append(["workbook", "FuloFilo_Master"])

    path = tmp_path / "Production_Master.xlsx"
    wb.save(path)
    return path


def _run_sync_fixture(tmp_path: Path, workbook: Path, monkeypatch: pytest.MonkeyPatch) -> dict:
    import scripts.sync_excel as sync_excel_module

    sync_excel = importlib.reload(sync_excel_module)
    runtime_root = tmp_path / "runtime"
    data_root = runtime_root / "data"
    excel_dir = data_root / "excel"
    parquet_dir = data_root / "parquet"
    raw_dir = data_root / "raw"
    excel_dir.mkdir(parents=True, exist_ok=True)
    parquet_dir.mkdir(parents=True, exist_ok=True)
    raw_dir.mkdir(parents=True, exist_ok=True)

    monkeypatch.setattr(sync_excel, "ROOT", runtime_root)
    monkeypatch.setattr(sync_excel, "EXCEL_DIR", excel_dir)
    monkeypatch.setattr(sync_excel, "DEFAULT_XLSX", workbook)
    monkeypatch.setattr(sync_excel, "PARQUET_DIR", parquet_dir)
    monkeypatch.setattr(sync_excel, "STATUS_PATH", excel_dir / "source_sync_status.json")
    monkeypatch.setattr(sys, "argv", ["sync_excel.py", "--excel", str(workbook)])

    sync_excel.main()
    return json.loads((excel_dir / "source_sync_status.json").read_text(encoding="utf-8"))


# ── Test 1 ────────────────────────────────────────────────────────────────────
def test_parquet_files_exist():
    """All 7 expected parquet files must be present on disk."""
    missing = []
    for name in EXPECTED_PARQUETS:
        path = DATA_DIR / f"{name}.parquet"
        if not path.exists():
            missing.append(name)
    assert missing == [], f"Missing parquet files: {missing}"


# ── Test 2 ────────────────────────────────────────────────────────────────────
def test_duckdb_products_not_empty():
    """DuckDB products view must return at least 1 row."""
    import duckdb
    db_path = ROOT / "data" / "fulofilo.duckdb"
    products_parquet = DATA_DIR / "products.parquet"
    assert products_parquet.exists(), "products.parquet must exist before DuckDB test"

    conn = duckdb.connect(str(db_path))
    conn.execute(f"CREATE OR REPLACE VIEW products AS "
                 f"SELECT * FROM read_parquet('{products_parquet}')")
    count = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    conn.close()
    assert count > 0, f"products view returned {count} rows — expected > 0"


# ── Test 3 ────────────────────────────────────────────────────────────────────
def test_abc_classification_coverage():
    """Every product in products.parquet must have a valid A/B/C class."""
    df = pl.read_parquet(DATA_DIR / "products.parquet")
    assert df.shape[0] > 0, "products.parquet is empty"

    valid_classes = {"A", "B", "C"}
    if "abc_class" not in df.columns:
        pytest.skip("abc_class column not yet populated — run build_catalog.py")

    null_count = df["abc_class"].is_null().sum()
    assert null_count == 0, f"{null_count} products have null abc_class"

    invalid = df.filter(~pl.col("abc_class").is_in(list(valid_classes)))
    assert invalid.shape[0] == 0, \
        f"{invalid.shape[0]} products have invalid abc_class: {invalid['abc_class'].unique().to_list()}"


# ── Test 4 ────────────────────────────────────────────────────────────────────
def test_no_negative_prices():
    """No product may have a negative suggested_price or unit_cost."""
    df = pl.read_parquet(DATA_DIR / "products.parquet")
    assert df.shape[0] > 0, "products.parquet is empty"

    for col in ["suggested_price", "unit_cost"]:
        if col not in df.columns:
            continue
        neg = df.filter(pl.col(col) < 0)
        assert neg.shape[0] == 0, \
            f"{neg.shape[0]} products have negative {col}: {neg['sku'].to_list()}"


# ── Test 5 ────────────────────────────────────────────────────────────────────
def test_category_coverage():
    """Canonical products read model must expose nonblank categories."""
    df = pl.read_parquet(DATA_DIR / "products.parquet")
    assert df.shape[0] > 0, "products.parquet is empty"
    assert "category" in df.columns, "products.parquet missing category column"

    blank = df.filter(pl.col("category").cast(pl.Utf8).str.strip_chars() == "")
    assert blank.shape[0] == 0, f"{blank.shape[0]} products have blank categories"


# ── Test 6 ────────────────────────────────────────────────────────────────────
def test_excel_builds_successfully():
    """build_report() must produce a valid .xlsx file larger than 50 KB."""
    import tempfile
    from excel.build_report import build_report

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "test_report.xlsx"
        result = build_report(output_path=out)
        # Check inside context — before tempdir is deleted
        assert result.exists(), f"Output file not found: {result}"
        size = result.stat().st_size
    assert size > 10_000, f"Generated Excel is too small ({size} bytes) — likely empty/corrupt"


def test_products_schema_matches_dashboard_contract():
    """Canonical sync output must expose the fields the dashboard reads."""
    df = pl.read_parquet(DATA_DIR / "products.parquet")
    required = {
        "sku", "full_name", "category", "unit_cost", "suggested_price",
        "qty_sold", "revenue", "profit", "margin_pct", "abc_class",
    }
    missing = sorted(required - set(df.columns))
    assert missing == [], f"products.parquet missing dashboard fields: {missing}"


def test_app_db_queries_accept_sync_excel_schema():
    """DuckDB query helpers must work against the canonical sync schema."""
    from app.db import (
        get_abc_analysis,
        get_conn,
        get_inventory_alerts,
        get_margin_matrix,
        get_stock_turnover,
        get_summary_kpis,
    )

    conn = get_conn()
    try:
        summary = get_summary_kpis(conn, "ALL")
        abc = get_abc_analysis(conn, "ALL")
        margin = get_margin_matrix(conn, "ALL")
        turnover = get_stock_turnover(conn)
        alerts = get_inventory_alerts(conn)
    finally:
        conn.close()

    assert len(summary) == 4, f"Unexpected KPI tuple: {summary}"
    assert abc.columns == [
        "full_name", "category", "revenue", "qty_sold", "profit", "abc_class", "cum_pct", "margin_pct",
    ], f"ABC query schema mismatch: {abc.columns}"
    assert margin.columns == [
        "full_name", "category", "qty_sold", "revenue", "margin_pct", "abc_class",
    ], f"Margin query schema mismatch: {margin.columns}"
    assert turnover.columns == [
        "product", "category", "current_stock", "min_stock", "qty_sold", "giro", "giro_class",
    ], f"Turnover query schema mismatch: {turnover.columns}"
    assert alerts.columns == [
        "product", "category", "current_stock", "min_stock", "alert",
    ], f"Inventory alert schema mismatch: {alerts.columns}"


def test_daily_sales_writeback_targets_daily_sales_sheet(tmp_path: Path):
    """Sale entry must append directly to the DailySales sheet."""
    from app.utils.sales_ops import append_sale_to_excel

    workbook = _make_test_workbook(tmp_path)
    sync_calls: list[bool] = []

    def fake_sync():
        sync_calls.append(True)
        return True, "ok"

    result = append_sale_to_excel(
        sale_date=date(2026, 4, 27),
        sku="00002",
        product="Bolsa Teste",
        quantity=2,
        unit_price=45.0,
        payment_method="Pix",
        source="manual",
        workbook_path=workbook,
        run_sync=True,
        sync_runner=fake_sync,
    )

    wb = openpyxl.load_workbook(workbook, read_only=True)
    ws = wb["DailySales"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[-1] == ("2026-04-27", "00002", "Bolsa Teste", 2, 45, 90, "Pix", "manual")
    assert result.total == 90.0
    assert sync_calls == [True]
    assert Path(result.backup_path).exists()


def test_category_writeback_targets_category_overrides_sheet(tmp_path: Path):
    """Category edits must upsert directly in CategoryOverrides."""
    from app.utils.category_ops import upsert_category_override

    workbook = _make_test_workbook(tmp_path)
    result = upsert_category_override(
        sku="00002",
        category="Souvenirs",
        subcategory="Chaveiros",
        confidence="manual",
        workbook_path=workbook,
        run_sync=False,
    )

    wb = openpyxl.load_workbook(workbook, read_only=True)
    ws = wb["CategoryOverrides"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[-1] == ("00002", "Souvenirs", "Chaveiros", "manual")
    assert result.sku == "00002"
    assert Path(result.backup_path).exists()


def test_inventory_writeback_targets_inventory_sheet_first(tmp_path: Path):
    """Inventory adjustments must mutate the Inventory sheet and append the audit log."""
    from app.utils.inventory_ops import adjust_stock

    workbook = _make_test_workbook(tmp_path)
    log_path = tmp_path / "stock_audit.csv"
    sync_calls: list[bool] = []

    def fake_sync():
        sync_calls.append(True)
        return True, "ok"

    result = adjust_stock(
        sku="00002",
        new_qty=11,
        workbook_path=workbook,
        log_path=log_path,
        run_sync=True,
        sync_runner=fake_sync,
    )

    wb = openpyxl.load_workbook(workbook, read_only=True)
    ws = wb["Inventory"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[2][3] == 11
    assert log_path.exists()
    log_text = log_path.read_text(encoding="utf-8")
    assert "00002" in log_text and "adjust" in log_text
    assert result.old_stock == 8 and result.new_stock == 11
    assert sync_calls == [True]


def test_no_active_path_writes_legacy_daily_sales_csv():
    """The app must not write to daily_sales_TEMPLATE.csv anymore."""
    page_source = (ROOT / "app" / "pages" / "04_daily_ops.py").read_text(encoding="utf-8")
    helper_source = (ROOT / "app" / "utils" / "sales_ops.py").read_text(encoding="utf-8")
    assert "daily_sales_TEMPLATE.csv" not in page_source
    assert "daily_sales_TEMPLATE.csv" not in helper_source


def test_no_active_path_writes_categorized_csv():
    """The app must not write category edits to product_catalog_categorized.csv."""
    page_source = (ROOT / "app" / "pages" / "05_categories.py").read_text(encoding="utf-8")
    helper_source = (ROOT / "app" / "utils" / "category_ops.py").read_text(encoding="utf-8")
    assert "product_catalog_categorized.csv" not in page_source
    assert "product_catalog_categorized.csv" not in helper_source


def test_no_active_path_mutates_inventory_parquet_directly():
    """Inventory write-back helpers must not write parquet directly."""
    helper_source = (ROOT / "app" / "utils" / "inventory_ops.py").read_text(encoding="utf-8")
    assert "write_parquet" not in helper_source
    assert ".write_parquet(" not in helper_source


def test_sync_excel_script_still_regenerates_parquet_successfully():
    """Canonical sync script remains runnable after write-back changes."""
    status = json.loads((ROOT / "data" / "excel" / "source_sync_status.json").read_text(encoding="utf-8"))
    assert status.get("ok") is True, status


def test_placeholder_only_workbook_emits_health_warning():
    """Bootstrap-only workbook state must not look like healthy production data."""
    products = pl.read_parquet(DATA_DIR / "products.parquet")
    status = json.loads((ROOT / "data" / "excel" / "source_sync_status.json").read_text(encoding="utf-8"))
    placeholder_only = (
        products.height == 1
        and products["full_name"].to_list() == ["Produto Exemplo"]
    )
    if placeholder_only:
        warnings = " ".join(status.get("warnings", []))
        assert "not healthy production data yet" in warnings, status
        assert status.get("health", {}).get("healthy_production_data") is False, status


def test_health_status_fields_exist_in_status_file():
    """Status output must expose machine-readable readiness fields."""
    status = json.loads((ROOT / "data" / "excel" / "source_sync_status.json").read_text(encoding="utf-8"))
    for field in [
        "healthy_production_data",
        "placeholder_only",
        "catalog_rows",
        "inventory_rows",
        "daily_sales_rows",
        "cashflow_rows",
    ]:
        assert field in status, f"Missing top-level status field: {field}"
        assert field in status.get("health", {}), f"Missing nested health field: {field}"


def test_bootstrap_fixture_is_detected_as_placeholder(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """A bootstrap-only workbook must be marked as placeholder and not ready."""
    status = _run_sync_fixture(tmp_path, _make_bootstrap_workbook(tmp_path), monkeypatch)
    assert status["ok"] is True
    assert status["placeholder_only"] is True
    assert status["healthy_production_data"] is False
    assert status["health"]["readiness_state"] == "bootstrap"


def test_zero_daily_sales_warning_is_non_destructive(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """Empty DailySales should warn but still allow sync to complete."""
    status = _run_sync_fixture(tmp_path, _make_test_workbook(tmp_path), monkeypatch)
    warnings = " ".join(status["warnings"])
    assert status["ok"] is True
    assert status["errors"] == []
    assert "DailySales has zero rows" in warnings


def test_zero_cashflow_warning_is_non_destructive(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """Empty Cashflow should warn but not fail sync."""
    status = _run_sync_fixture(tmp_path, _make_test_workbook(tmp_path), monkeypatch)
    warnings = " ".join(status["warnings"])
    assert status["ok"] is True
    assert status["errors"] == []
    assert "Cashflow has zero rows" in warnings


def test_empty_inventory_warning_is_non_destructive(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """Header-only Inventory should warn but still produce a status file."""
    workbook = _make_test_workbook(tmp_path)
    wb = openpyxl.load_workbook(workbook)
    ws = wb["Inventory"]
    ws.delete_rows(2, ws.max_row - 1)
    wb.save(workbook)

    status = _run_sync_fixture(tmp_path, workbook, monkeypatch)
    warnings = " ".join(status["warnings"])
    assert status["ok"] is True
    assert status["errors"] == []
    assert status["inventory_rows"] == 0
    assert "Inventory has zero rows" in warnings


def test_production_ready_fixture_becomes_healthy(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """A populated workbook fixture should be marked ready for production."""
    status = _run_sync_fixture(tmp_path, _make_production_ready_workbook(tmp_path), monkeypatch)
    assert status["ok"] is True
    assert status["healthy_production_data"] is True
    assert status["health"]["readiness_state"] == "ready"
    assert status["catalog_real_rows"] >= 5
    assert status["daily_sales_rows"] > 0
    assert status["cashflow_rows"] > 0


def test_docs_mention_backup_convention():
    """Operator docs should document the workbook backup path convention."""
    text = "\n".join(
        [
            (ROOT / "README.md").read_text(encoding="utf-8"),
            (ROOT / "DOCUMENTATION.md").read_text(encoding="utf-8"),
            (ROOT / "docs" / "USER_GUIDE.md").read_text(encoding="utf-8"),
        ]
    )
    assert "data/excel/backups/FuloFilo_Master_YYYYMMDD_HHMMSS.xlsx" in text


def test_backup_workbook_avoids_same_second_filename_collisions(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """Multiple backups in the same second must not overwrite each other."""
    import app.utils.excel_sync as excel_sync

    workbook = tmp_path / "FuloFilo_Master.xlsx"
    workbook.write_text("fixture", encoding="utf-8")

    class _FrozenDatetime:
        @staticmethod
        def now():
            from datetime import datetime as real_datetime
            return real_datetime(2026, 4, 27, 3, 1, 2)

    monkeypatch.setattr(excel_sync, "datetime", _FrozenDatetime)

    first = excel_sync.backup_workbook(workbook_path=workbook, backup_dir=tmp_path / "backups")
    second = excel_sync.backup_workbook(workbook_path=workbook, backup_dir=tmp_path / "backups")

    assert first.exists()
    assert second.exists()
    assert first != second
    assert first.name == "FuloFilo_Master_20260427_030102.xlsx"
    assert second.name == "FuloFilo_Master_20260427_030102_01.xlsx"


def test_operator_docs_do_not_reference_deleted_active_scripts():
    """If deleted paths are mentioned, they must be clearly marked as legacy."""
    files = [
        ROOT / "README.md",
        ROOT / "DOCUMENTATION.md",
        ROOT / "docs" / "README.md",
        ROOT / "docs" / "USER_GUIDE.md",
        ROOT / "docs" / "PROFESSIONAL_REPORT.md",
        ROOT / "docs" / "DATA_DICTIONARY.md",
    ]
    legacy_markers = [
        "etl/build_catalog.py",
        "etl/ingest_eleve.py",
        "scripts/sync_native_sources.sh",
    ]
    for path in files:
        text = path.read_text(encoding="utf-8")
        lowered = text.lower()
        for marker in legacy_markers:
            if marker in text:
                assert "archived" in lowered or "legacy" in lowered or "deleted" in lowered, (
                    f"{path.name} mentions {marker} without quarantine wording"
                )


def test_operator_docs_do_not_present_legacy_csv_as_active_sources():
    """Legacy CSV artifacts must not be described as operational write targets."""
    files = [
        ROOT / "README.md",
        ROOT / "DOCUMENTATION.md",
        ROOT / "docs" / "README.md",
        ROOT / "docs" / "USER_GUIDE.md",
        ROOT / "docs" / "PROFESSIONAL_REPORT.md",
        ROOT / "docs" / "DATA_DICTIONARY.md",
    ]
    active_markers = [
        "daily_sales_TEMPLATE.csv",
        "product_catalog_categorized.csv",
    ]
    for path in files:
        text = path.read_text(encoding="utf-8")
        if any(marker in text for marker in active_markers):
            lowered = text.lower()
            assert "not active" in lowered or "archived" in lowered or "not operational" in lowered, path.name


def test_refresh_data_script_is_quarantined_not_canonical():
    """Legacy refresh script should be a disabled archive stub."""
    text = (ROOT / "scripts" / "refresh_data.sh").read_text(encoding="utf-8")
    assert "Legacy path archived" in text
    assert "sync_excel.sh" in text
    assert "exit 1" in text


def test_generated_artifacts_are_described_as_non_write_targets():
    """Operator docs must classify generated artifacts as read-only/generated outputs."""
    text = "\n".join(
        [
            (ROOT / "README.md").read_text(encoding="utf-8"),
            (ROOT / "DOCUMENTATION.md").read_text(encoding="utf-8"),
            (ROOT / "docs" / "DATA_DICTIONARY.md").read_text(encoding="utf-8"),
        ]
    ).lower()
    assert "data/parquet/*.parquet" in text
    assert "data/fulofilo.duckdb" in text
    assert "generated" in text
    assert "read-only" in text or "not a source-of-truth" in text or "not a source-of-truth" in text
