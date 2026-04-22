#!/usr/bin/env python3
"""
PDF daily sales -> CSV/XLSX pipeline

Expected PDF pattern:
"venda 05 unidades de ' CAMISETA BASICA ' da estampa ' Cabedelo_Peito ', tamanho ' M '"

Outputs:
- sales_entries.csv
- sales_entries.xlsx
- parse_log.txt
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime, date
from pathlib import Path
from typing import Iterable, List, Optional

try:
    import fitz  # PyMuPDF
except Exception as exc:
    raise SystemExit("Missing dependency: pymupdf (PyMuPDF). Install with: python3 -m pip install pymupdf openpyxl") from exc

try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception as exc:
    raise SystemExit("Missing dependency: openpyxl. Install with: python3 -m pip install openpyxl") from exc


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

CATEGORY_MAP = {
    "BASICA": "BÁSICAS",
    "BASICA": "BÁSICAS",
    "BASICAS": "BÁSICAS",
    "REGATA": "REGATAS",
    "REGATAS": "REGATAS",
    "INFANTIL": "INFANTIS",
    "INFANTIS": "INFANTIS",
    "BABY": "BABY",
}

DATE_PATTERNS = [
    (re.compile(r"(?P<y>20\d{2})[-_](?P<m>\d{2})[-_](?P<d>\d{2})"), "ymd"),
    (re.compile(r"(?P<d>\d{2})[-_](?P<m>\d{2})[-_](?P<y>20\d{2})"), "dmy"),
    (re.compile(r"(?P<y>20\d{2})(?P<m>\d{2})(?P<d>\d{2})"), "ymd"),
]

SALE_REGEX = re.compile(
    r"venda\s+(?P<units>\d{1,3})\s+unidades?\s+de\s+'?\s*(?P<product>CAMISETA\s+[A-ZÀ-Ú_ ]+?)\s*'?\s+da\s+estampa\s+'?\s*(?P<print>[A-ZÀ-Ú0-9_ ]+?)\s*'?\s*,?\s*tamanho\s+'?\s*(?P<size>[A-Z0-9]+)\s*'?",
    re.IGNORECASE,
)

@dataclass
class SaleEntry:
    sale_date: str
    source_file: str
    source_path: str
    page: int
    category: str
    product_label: str
    print_name: str
    size: str
    units: int
    notes: str = ""


def normalize_spaces(text: str) -> str:
    text = text.replace("‘", "'").replace("’", "'").replace("`", "'").replace("´", "'")
    text = text.replace("_", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", " ", text)
    return text.strip()


def ascii_fold(text: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")


def normalize_category(product_label: str) -> str:
    folded = ascii_fold(product_label.upper())
    folded = re.sub(r"\s+", " ", folded).strip()
    folded = folded.replace("CAMISETA ", "")
    for key, value in CATEGORY_MAP.items():
        if key in folded:
            return value
    return "UNKNOWN"


def normalize_product(product_label: str) -> str:
    return re.sub(r"\s+", " ", product_label.upper()).strip().replace("_", " ")


def normalize_print(print_name: str) -> str:
    clean = re.sub(r"\s+", " ", print_name.upper()).strip().replace("_", " ")
    return clean


def detect_date(pdf_path: Path, explicit_default: Optional[str] = None) -> str:
    if explicit_default:
        return explicit_default

    name = pdf_path.stem
    for pattern, mode in DATE_PATTERNS:
        match = pattern.search(name)
        if match:
            y, m, d = int(match.group("y")), int(match.group("m")), int(match.group("d"))
            return date(y, m, d).isoformat()

    mtime = datetime.fromtimestamp(pdf_path.stat().st_mtime)
    return mtime.date().isoformat()


def extract_text(pdf_path: Path) -> List[str]:
    pages = []
    with fitz.open(pdf_path) as doc:
        for page in doc:
            pages.append(page.get_text("text"))
    return pages


def parse_pdf(pdf_path: Path, default_date: Optional[str] = None) -> List[SaleEntry]:
    sale_date = detect_date(pdf_path, default_date)
    entries: List[SaleEntry] = []

    for page_number, raw_text in enumerate(extract_text(pdf_path), start=1):
        text = normalize_spaces(raw_text)
        for match in SALE_REGEX.finditer(text):
            product = normalize_product(match.group("product"))
            print_name = normalize_print(match.group("print"))
            size = match.group("size").upper().strip()
            units = int(match.group("units"))
            category = normalize_category(product)

            entries.append(
                SaleEntry(
                    sale_date=sale_date,
                    source_file=pdf_path.name,
                    source_path=str(pdf_path),
                    page=page_number,
                    category=category,
                    product_label=product,
                    print_name=print_name,
                    size=size,
                    units=units,
                    notes="Extracted from sale annotation",
                )
            )
    return entries


def write_csv(entries: List[SaleEntry], csv_path: Path) -> None:
    fieldnames = list(asdict(entries[0]).keys()) if entries else [
        "sale_date", "source_file", "source_path", "page", "category",
        "product_label", "print_name", "size", "units", "notes"
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for entry in entries:
            writer.writerow(asdict(entry))


def autosize(ws) -> None:
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            val = "" if cell.value is None else str(cell.value)
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), min(len(val) + 2, 36))
            cell.alignment = Alignment(vertical="center")
            cell.border = THIN_BORDER
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header(ws, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def add_table(ws, start_cell: str, end_cell: str, table_name: str) -> None:
    tab = Table(displayName=table_name, ref=f"{start_cell}:{end_cell}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def create_excel(entries: List[SaleEntry], xlsx_path: Path) -> None:
    wb = Workbook()
    raw = wb.active
    raw.title = "Raw_Entries"

    headers = ["sale_date", "source_file", "source_path", "page", "category",
               "product_label", "print_name", "size", "units", "notes"]
    raw.append(headers)
    for entry in entries:
        raw.append([getattr(entry, h) for h in headers])

    style_header(raw, 1)
    if len(entries) > 0:
        add_table(raw, "A1", f"J{len(entries)+1}", "RawEntriesTable")
    raw.freeze_panes = "A2"
    autosize(raw)

    # Daily summary
    daily = wb.create_sheet("Daily_Summary")
    daily.append(["sale_date", "total_units", "transactions"])
    by_day = defaultdict(lambda: {"units": 0, "transactions": 0})
    for entry in entries:
        by_day[entry.sale_date]["units"] += entry.units
        by_day[entry.sale_date]["transactions"] += 1
    for sale_date in sorted(by_day):
        daily.append([sale_date, by_day[sale_date]["units"], by_day[sale_date]["transactions"]])
    style_header(daily, 1)
    if len(by_day) > 0:
        add_table(daily, "A1", f"C{len(by_day)+1}", "DailySummaryTable")
    autosize(daily)

    # Print summary
    pr = wb.create_sheet("Print_Summary")
    pr.append(["print_name", "total_units"])
    by_print = defaultdict(int)
    for entry in entries:
        by_print[entry.print_name] += entry.units
    for print_name, units in sorted(by_print.items(), key=lambda kv: (-kv[1], kv[0])):
        pr.append([print_name, units])
    style_header(pr, 1)
    if len(by_print) > 0:
        add_table(pr, "A1", f"B{len(by_print)+1}", "PrintSummaryTable")
    autosize(pr)

    # Charts
    if len(by_day) > 0:
        line = LineChart()
        line.title = "Units per Day"
        line.y_axis.title = "Units"
        line.x_axis.title = "Date"
        data = Reference(daily, min_col=2, min_row=1, max_row=len(by_day)+1)
        cats = Reference(daily, min_col=1, min_row=2, max_row=len(by_day)+1)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height = 8
        line.width = 16
        daily.add_chart(line, "E2")

    if len(by_print) > 0:
        bar = BarChart()
        bar.title = "Top Prints"
        bar.y_axis.title = "Units"
        data = Reference(pr, min_col=2, min_row=1, max_row=min(len(by_print)+1, 11))
        cats = Reference(pr, min_col=1, min_row=2, max_row=min(len(by_print)+1, 11))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 8
        bar.width = 16
        pr.add_chart(bar, "D2")

    wb.save(xlsx_path)


def discover_pdfs(input_dir: Path, recursive: bool) -> List[Path]:
    pattern = "**/*.pdf" if recursive else "*.pdf"
    return sorted(input_dir.glob(pattern))


def write_log(log_path: Path, parsed_files: List[Path], entries: List[SaleEntry], skipped_files: List[str]) -> None:
    with log_path.open("w", encoding="utf-8") as handle:
        handle.write("PDF to sales pipeline log\n")
        handle.write("=" * 30 + "\n")
        handle.write(f"Parsed PDFs: {len(parsed_files)}\n")
        handle.write(f"Entries extracted: {len(entries)}\n")
        handle.write("\nFiles:\n")
        for item in parsed_files:
            handle.write(f"- {item}\n")
        if skipped_files:
            handle.write("\nSkipped / no matches:\n")
            for item in skipped_files:
                handle.write(f"- {item}\n")


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert daily sales PDFs into CSV and Excel.")
    parser.add_argument("--input-dir", required=True, help="Folder containing PDF files.")
    parser.add_argument("--output-dir", required=True, help="Folder where CSV/XLSX outputs will be written.")
    parser.add_argument("--recursive", action="store_true", help="Search subfolders too.")
    parser.add_argument("--default-date", help="Fallback sale date in YYYY-MM-DD. Overrides filename/mtime.")
    args = parser.parse_args()

    input_dir = Path(args.input_dir).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    if not input_dir.exists():
        raise SystemExit(f"Input directory not found: {input_dir}")

    pdfs = discover_pdfs(input_dir, recursive=args.recursive)
    if not pdfs:
        raise SystemExit(f"No PDF files found in: {input_dir}")

    all_entries: List[SaleEntry] = []
    parsed_files: List[Path] = []
    skipped_files: List[str] = []

    for pdf_path in pdfs:
        try:
            entries = parse_pdf(pdf_path, default_date=args.default_date)
            if entries:
                all_entries.extend(entries)
                parsed_files.append(pdf_path)
            else:
                skipped_files.append(f"{pdf_path} (no sale annotations matched)")
        except Exception as exc:
            skipped_files.append(f"{pdf_path} (error: {exc})")

    csv_path = output_dir / "sales_entries.csv"
    xlsx_path = output_dir / "sales_entries.xlsx"
    log_path = output_dir / "parse_log.txt"

    write_csv(all_entries, csv_path)
    create_excel(all_entries, xlsx_path)
    write_log(log_path, parsed_files, all_entries, skipped_files)

    print(f"Done. CSV: {csv_path}")
    print(f"Done. XLSX: {xlsx_path}")
    print(f"Log: {log_path}")
    print(f"Entries extracted: {len(all_entries)}")
    if skipped_files:
        print("Some PDFs were skipped. See parse_log.txt")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
