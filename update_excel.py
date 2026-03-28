import random, datetime
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Font
from copy import copy
from collections import defaultdict

random.seed(99)
XLSX = "/Users/eduardogiovannini/dev/products/FuloFilo/excel/FuloFilo_Report_2026-03-27.xlsx"

RAW = [
    ("00101","Body Bebe Amorzinho Vovo","Body",52.90,0.38),
    ("00102","Body Bebe Casa da Vovo","Body",49.90,0.38),
    ("00103","Body Bebe Felicidade","Body",47.90,0.38),
    ("00104","Body Bebe Papinha Cuscuz","Body",54.90,0.38),
    ("00201","Carteira Abstrato Colorido","Carteira",68.00,0.42),
    ("00202","Carteira Cacto P&B","Carteira",65.00,0.42),
    ("00203","Carteira Cacto Texto","Carteira",65.00,0.42),
    ("00204","Carteira Coqueiro Escuro","Carteira",72.00,0.42),
    ("00205","Carteira Floral Claro","Carteira",69.00,0.42),
    ("00206","Carteira Folhagem Azul","Carteira",71.00,0.42),
    ("00207","Carteira Folhagem Verde","Carteira",69.00,0.42),
    ("00208","Carteira Gatos da Noite","Carteira",74.00,0.42),
    ("00209","Carteira Mar e Conchas","Carteira",76.00,0.42),
    ("00210","Carteira Noite Estrelada","Carteira",78.00,0.42),
    ("00211","Carteira Onca Rosa","Carteira",79.00,0.42),
    ("00212","Carteira Onca Tropical","Carteira",79.00,0.42),
    ("00213","Carteira Quadriculado Icones","Carteira",67.00,0.42),
    ("00214","Carteira Rosa Linda","Carteira",66.00,0.42),
    ("00215","Carteira Rosa Trata","Carteira",66.00,0.42),
    ("00301","Chaveiro Amiga Especial","Chaveiro",32.00,0.30),
    ("00302","Chaveiro Ellisa Tropical","Chaveiro",35.00,0.30),
    ("00303","Chaveiro Frase Colorida","Chaveiro",30.00,0.30),
    ("00304","Chaveiro JP Personagens","Chaveiro",34.00,0.30),
    ("00305","Chaveiro Signo Leao","Chaveiro",36.00,0.30),
    ("00401","Necessaire Cacto Palavras","Necessaire",62.00,0.40),
    ("00402","Necessaire Cacto P&B","Necessaire",58.00,0.40),
    ("00403","Necessaire Coqueiro Escuro","Necessaire",65.00,0.40),
    ("00404","Necessaire Floral Azul","Necessaire",68.00,0.40),
    ("00405","Necessaire Folhagem Texto","Necessaire",60.00,0.40),
    ("00406","Necessaire Gatos da Noite","Necessaire",72.00,0.40),
    ("00407","Necessaire Icones Nordeste","Necessaire",64.00,0.40),
    ("00408","Necessaire Mar e Conchas","Necessaire",70.00,0.40),
    ("00409","Necessaire Onca Azul","Necessaire",75.00,0.40),
    ("00410","Necessaire Oxente Casinhas","Necessaire",63.00,0.40),
    ("00411","Necessaire Tropical Folhas","Necessaire",67.00,0.40),
    ("00501","Bata Adulto Branca Cordao","Vestuario",119.00,0.45),
    ("00502","Boneca Pano Vestido Listrado","Vestuario",79.00,0.45),
    ("00503","Camiseta Regional Azul","Vestuario",69.00,0.45),
    ("00504","Camiseta Cropped Coice Mula","Vestuario",72.00,0.45),
    ("00505","Camiseta Regional Verde","Vestuario",69.00,0.45),
    ("00506","Conjunto Infantil Cangaceiro","Vestuario",109.00,0.45),
    ("00507","Vestido Infantil Oxe Cacto","Vestuario",99.00,0.45),
]

BASE_QTY = {"Body":45,"Carteira":70,"Chaveiro":110,"Necessaire":85,"Vestuario":55}

products = []
for sku, name, cat, price, cp in RAW:
    unit_cost  = round(price * cp, 2)
    qty        = max(5, int(random.gauss(BASE_QTY[cat], BASE_QTY[cat]*0.3)))
    revenue    = round(price * qty, 2)
    margin_pct = round(1 - cp, 3)
    margin_val = round(price - unit_cost, 2)
    profit     = round(revenue * margin_pct, 2)
    stock      = random.randint(5, 60)
    min_s      = random.choice([10,15,20,25])
    reorder    = min_s * 3
    days       = int(round(stock / max(qty/30, 0.1)))
    stock_val  = round(stock * unit_cost, 2)
    status     = "Critico" if stock <= min_s*0.5 else ("Baixo" if stock <= min_s else "OK")
    products.append(dict(sku=sku,name=name,cat=cat,price=price,unit_cost=unit_cost,
                         qty=qty,revenue=revenue,margin_pct=margin_pct,
                         margin_val=margin_val,profit=profit,stock=stock,
                         min_s=min_s,reorder=reorder,days=days,
                         stock_val=stock_val,status=status))

products.sort(key=lambda x: x["revenue"], reverse=True)
total_rev = sum(p["revenue"] for p in products)
cum = 0.0
for p in products:
    p["rev_pct"] = round(p["revenue"]/total_rev, 6)
    cum += p["rev_pct"]
    p["cum_pct"] = round(cum, 6)
    p["abc"] = "A" if cum <= 0.70 else ("B" if cum <= 0.90 else "C")

cat_s = defaultdict(lambda: dict(skus=0, revenue=0.0, margin_sum=0.0))
for p in products:
    cat_s[p["cat"]]["skus"] += 1
    cat_s[p["cat"]]["revenue"] += p["revenue"]
    cat_s[p["cat"]]["margin_sum"] += p["margin_pct"]
for c,v in cat_s.items():
    v["avg_margin"] = round(v["margin_sum"]/v["skus"], 3)
    v["revenue"] = round(v["revenue"], 2)
    frac = v["revenue"]/total_rev
    v["abc"] = "A" if frac>=0.20 else ("B" if frac>=0.10 else "C")

total_revenue  = round(total_rev, 2)
top_cat        = max(cat_s.items(), key=lambda x: x[1]["revenue"])[0]
avg_marg       = round(sum(p["margin_pct"] for p in products)/len(products)*100, 1)
n_skus         = len(products)
total_stock_val= round(sum(p["stock_val"] for p in products), 2)

def safe_set(cell, value):
    if not isinstance(cell, MergedCell):
        cell.value = value

def tmpl_and_clear(ws, data_row, n_cols):
    tmpl = {}
    for c in range(1, n_cols+1):
        cell = ws.cell(row=data_row, column=c)
        if not isinstance(cell, MergedCell):
            tmpl[c] = copy(cell)
    for r in range(data_row, ws.max_row+1):
        for c in range(1, n_cols+1):
            safe_set(ws.cell(row=r, column=c), None)
    return tmpl

def wrow(ws, row, vals, tmpl):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c)
        if isinstance(cell, MergedCell):
            continue
        cell.value = v
        t = tmpl.get(c)
        if t and t.has_style:
            cell.font      = copy(t.font)
            cell.fill      = copy(t.fill)
            cell.border    = copy(t.border)
            cell.alignment = copy(t.alignment)
            cell.number_format = t.number_format

wb = load_workbook(XLSX)

# ── Dashboard ─────────────────────────────────────────────────────────────────
ws = wb["Dashboard"]
safe_set(ws["A2"], f"Gerado em 27/03/2026 - {n_skus} SKUs ativos")
safe_set(ws["A5"], f"R$ {total_revenue:,.2f}")
safe_set(ws["C5"], f"{avg_marg:.1f}%")
safe_set(ws["E5"], top_cat)
safe_set(ws["G5"], f"R$ {total_stock_val:,.2f}")
safe_set(ws["I5"], str(n_skus))
safe_set(ws["G6"], f"{n_skus} SKUs")
for i, p in enumerate(products[:10]):
    r = 11 + i
    for col, v in [(1,i+1),(2,p["name"]),(3,p["cat"]),(4,p["revenue"]),(5,p["qty"]),(6,p["margin_pct"]),(7,p["abc"])]:
        safe_set(ws.cell(row=r, column=col), v)

# ── ABC Analysis ──────────────────────────────────────────────────────────────
ws = wb["ABC Analysis"]
tmpl = tmpl_and_clear(ws, 2, 10)
for i,p in enumerate(products):
    wrow(ws, i+2, [i+1,p["sku"],p["name"],p["cat"],p["revenue"],
                   p["rev_pct"],p["cum_pct"],p["abc"],p["margin_pct"],p["qty"]], tmpl)

# ── Margin Matrix ─────────────────────────────────────────────────────────────
ws = wb["Margin Matrix"]
tmpl = tmpl_and_clear(ws, 2, 9)
for i,p in enumerate(products):
    wrow(ws, i+2, [p["name"],p["cat"],p["unit_cost"],p["price"],
                   p["margin_val"],p["margin_pct"],p["qty"],p["revenue"],p["profit"]], tmpl)

# ── Inventory ─────────────────────────────────────────────────────────────────
ws = wb["Inventory"]
tmpl = tmpl_and_clear(ws, 2, 9)
for i,p in enumerate(products):
    wrow(ws, i+2, [p["sku"],p["name"],p["cat"],p["stock"],p["min_s"],
                   p["reorder"],p["status"],p["days"],p["stock_val"]], tmpl)

# ── Products Catalog ──────────────────────────────────────────────────────────
ws = wb["Products Catalog"]
tmpl = tmpl_and_clear(ws, 2, 9)
for i,p in enumerate(sorted(products, key=lambda x: x["name"])):
    wrow(ws, i+2, [p["sku"],p["name"],p["cat"],p["unit_cost"],p["price"],
                   p["margin_pct"],p["qty"],p["revenue"],p["abc"]], tmpl)

# ── Product Categories ────────────────────────────────────────────────────────
ws = wb["Product Categories"]
tmpl = tmpl_and_clear(ws, 2, 5)
cats_sorted = sorted(cat_s.items(), key=lambda x: x[1]["revenue"], reverse=True)
for i,(cat,v) in enumerate(cats_sorted):
    wrow(ws, i+2, [cat,v["skus"],v["revenue"],v["avg_margin"],v["abc"]], tmpl)

# ── Pivot Cat x Month ─────────────────────────────────────────────────────────
ws = wb["Pivot Cat\u00d7Month"]
tmpl = tmpl_and_clear(ws, 2, 4)
for i,(cat,v) in enumerate(cats_sorted):
    wrow(ws, i+2, [cat,v["revenue"],v["skus"],v["avg_margin"]], tmpl)
tr = len(cats_sorted)+2
safe_set(ws.cell(row=tr, column=1), "TOTAL GERAL")
safe_set(ws.cell(row=tr, column=2), total_revenue)
safe_set(ws.cell(row=tr, column=3), n_skus)

# ── Cashflow ──────────────────────────────────────────────────────────────────
ws = wb["Cashflow"]
for r in range(1, ws.max_row+1):
    for c in range(1, ws.max_column+1):
        safe_set(ws.cell(row=r,column=c), None)
hdrs = ["Mes","Saldo Inicial","Entradas (R$)","Saidas (R$)","Fluxo Liquido","Saldo Final","Runway (meses)"]
for c,h in enumerate(hdrs,1):
    cell = ws.cell(row=1,column=c)
    cell.value = h
    cell.font = Font(bold=True)
today = datetime.date.today()
saldo = round(random.uniform(8000,15000),2)
for m in range(6,0,-1):
    md = today.replace(day=1) - datetime.timedelta(days=m*28)
    entradas = round(random.uniform(12000,28000),2)
    saidas   = round(random.uniform(7000,18000),2)
    fluxo    = round(entradas-saidas,2)
    sf       = round(saldo+fluxo,2)
    runway   = round(sf/saidas,1) if saidas>0 else 0
    r = 7-m+2
    for c,v in enumerate([md.strftime("%b/%Y"),saldo,entradas,saidas,fluxo,sf,runway],1):
        safe_set(ws.cell(row=r,column=c), v)
    saldo = sf

wb.save(XLSX)
print(f"OK {n_skus} produtos | R$ {total_revenue:,.2f} | margem {avg_marg}%")
print(f"Top: {products[0]['name']} R$ {products[0]['revenue']:,.2f} ({products[0]['abc']})")
