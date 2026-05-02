from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
import polars as pl
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.excel_sync import MASTER_PATH, backup_workbook, run_canonical_sync
from app.utils.workbook_lock import locked_workbook

ROOT = Path(__file__).resolve().parent.parent.parent
INV_PATH = ROOT / "data" / "parquet" / "inventory.parquet"
LOG_PATH = ROOT / "data" / "logs" / "stock_audit.csv"

SHEET_INVENTORY = "Inventory"
ALLOW_NEGATIVE_STOCK = False

_COL_SKU = 1
_COL_PRODUCT = 2
_COL_CURRENT_STOCK = 4
_AUDIT_HEADERS = ["timestamp", "slug", "action", "qty_before", "qty_after", "delta"]


@dataclass
class InventoryWriteResult:
    sku: str
    product: str
    old_stock: int
    new_stock: int
    delta: int
    workbook_path: str
    backup_path: str | None = None
    sync_output: str = ""


def _norm_sku(value: object) -> str:
    if value is None:
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    try:
        return str(int(float(raw))).zfill(5)
    except ValueError:
        return raw


def load_inventory() -> pl.DataFrame:
    return pl.read_parquet(INV_PATH) if INV_PATH.exists() else pl.DataFrame()


def _append_audit_log(
    sku: str,
    action: str,
    qty_before: int,
    qty_after: int,
    log_path: Path = LOG_PATH,
) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not log_path.exists() or log_path.stat().st_size == 0
    with log_path.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=_AUDIT_HEADERS)
        if write_header:
            writer.writeheader()
        writer.writerow(
            {
                "timestamp": datetime.now().isoformat(timespec="seconds"),
                "slug": sku,
                "action": action,
                "qty_before": qty_before,
                "qty_after": qty_after,
                "delta": qty_after - qty_before,
            }
        )


def _require_inventory_sheet(wb: openpyxl.Workbook) -> Worksheet:
    if SHEET_INVENTORY not in wb.sheetnames:
        raise ValueError(f"Required sheet missing from workbook: {SHEET_INVENTORY}")
    return wb[SHEET_INVENTORY]


def _find_inventory_row(ws: Worksheet, sku: str) -> tuple[int, str, int]:
    sku_norm = _norm_sku(sku)
    for row_idx in range(2, ws.max_row + 1):
        if _norm_sku(ws.cell(row_idx, _COL_SKU).value) != sku_norm:
            continue
        product = str(ws.cell(row_idx, _COL_PRODUCT).value or "").strip()
        current_stock = int(float(ws.cell(row_idx, _COL_CURRENT_STOCK).value or 0))
        return row_idx, product, current_stock
    raise ValueError(f"SKU not found in Inventory: {sku_norm}")


def _write_inventory_stock(
    sku: str,
    new_qty: int,
    action: str,
    workbook_path: Path = MASTER_PATH,
    log_path: Path = LOG_PATH,
    run_sync: bool = True,
    create_backup: bool = True,
    sync_runner=run_canonical_sync,
) -> InventoryWriteResult:
    sku_norm = _norm_sku(sku)
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    if new_qty < 0 and not ALLOW_NEGATIVE_STOCK:
        raise ValueError("Negative stock is not allowed by the current configuration.")

    backup_path = None
    with locked_workbook(workbook_path, owner="streamlit_inventory_ops"):
        backup_path = backup_workbook(workbook_path) if create_backup else None
        wb = openpyxl.load_workbook(workbook_path)
        ws = _require_inventory_sheet(wb)
        row_idx, product, old_stock = _find_inventory_row(ws, sku_norm)
        ws.cell(row_idx, _COL_CURRENT_STOCK, int(new_qty))
        wb.save(workbook_path)

    _append_audit_log(sku_norm, action, old_stock, int(new_qty), log_path=log_path)

    sync_output = ""
    if run_sync:
        ok, sync_output = sync_runner()
        if not ok:
            raise RuntimeError(sync_output or "Canonical sync failed after Inventory write-back.")

    return InventoryWriteResult(
        sku=sku_norm,
        product=product,
        old_stock=old_stock,
        new_stock=int(new_qty),
        delta=int(new_qty) - old_stock,
        workbook_path=str(workbook_path),
        backup_path=str(backup_path) if backup_path else None,
        sync_output=sync_output,
    )


def adjust_stock(
    sku: str,
    new_qty: int,
    workbook_path: Path = MASTER_PATH,
    log_path: Path = LOG_PATH,
    run_sync: bool = True,
    create_backup: bool = True,
    sync_runner=run_canonical_sync,
) -> InventoryWriteResult:
    """Set current stock directly in the canonical Inventory worksheet, then sync."""
    return _write_inventory_stock(
        sku=sku,
        new_qty=int(new_qty),
        action="adjust",
        workbook_path=workbook_path,
        log_path=log_path,
        run_sync=run_sync,
        create_backup=create_backup,
        sync_runner=sync_runner,
    )


def decrement_stock(
    sku: str,
    qty: int,
    workbook_path: Path = MASTER_PATH,
    log_path: Path = LOG_PATH,
    run_sync: bool = True,
    create_backup: bool = True,
    sync_runner=run_canonical_sync,
) -> InventoryWriteResult:
    """Decrement stock in the canonical Inventory worksheet, then sync."""
    if qty <= 0:
        raise ValueError("Quantity must be greater than zero.")

    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    ws = _require_inventory_sheet(wb)
    _, _, old_stock = _find_inventory_row(ws, sku)
    wb.close()
    new_qty = old_stock - int(qty)
    if not ALLOW_NEGATIVE_STOCK:
        new_qty = max(0, new_qty)

    return _write_inventory_stock(
        sku=sku,
        new_qty=new_qty,
        action="decrement",
        workbook_path=workbook_path,
        log_path=log_path,
        run_sync=run_sync,
        create_backup=create_backup,
        sync_runner=sync_runner,
    )
