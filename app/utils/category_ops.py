from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.excel_sync import MASTER_PATH, backup_workbook, run_canonical_sync

SHEET_CATALOG = "Catalog"
SHEET_OVERRIDES = "CategoryOverrides"


@dataclass
class CategoryOverrideResult:
    sku: str
    category: str
    subcategory: str
    confidence: str
    workbook_path: str
    backup_path: str | None = None
    sync_output: str = ""


def _norm_sku(value: object) -> str:
    if value is None:
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    try:
        return str(int(float(raw))).zfill(5)
    except ValueError:
        return raw


def _require_sheet(wb: openpyxl.Workbook, name: str) -> Worksheet:
    if name not in wb.sheetnames:
        raise ValueError(f"Required sheet missing from workbook: {name}")
    return wb[name]


def _catalog_skus(ws: Worksheet) -> set[str]:
    return {
        _norm_sku(row[0] if row else None)
        for row in ws.iter_rows(min_row=2, values_only=True)
        if _norm_sku(row[0] if row else None)
    }


def upsert_category_override(
    sku: str,
    category: str,
    subcategory: str,
    confidence: str = "manual",
    workbook_path: Path = MASTER_PATH,
    run_sync: bool = True,
    create_backup: bool = True,
    sync_runner=run_canonical_sync,
) -> CategoryOverrideResult:
    """Upsert a manual category override in the canonical workbook, then sync."""
    sku_norm = _norm_sku(sku)
    if not sku_norm:
        raise ValueError("SKU is required.")
    if not str(category).strip():
        raise ValueError("Category is required.")

    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    backup_path = backup_workbook(workbook_path) if create_backup else None
    wb = openpyxl.load_workbook(workbook_path)
    catalog_ws = _require_sheet(wb, SHEET_CATALOG)
    overrides_ws = _require_sheet(wb, SHEET_OVERRIDES)

    if sku_norm not in _catalog_skus(catalog_ws):
        raise ValueError(f"SKU not found in Catalog: {sku_norm}")

    target_row = None
    for row_idx in range(2, overrides_ws.max_row + 1):
        if _norm_sku(overrides_ws.cell(row_idx, 1).value) == sku_norm:
            target_row = row_idx
            break
    if target_row is None:
        target_row = overrides_ws.max_row + 1

    overrides_ws.cell(target_row, 1, sku_norm)
    overrides_ws.cell(target_row, 2, str(category).strip())
    overrides_ws.cell(target_row, 3, str(subcategory).strip())
    overrides_ws.cell(target_row, 4, str(confidence).strip() or "manual")
    wb.save(workbook_path)

    sync_output = ""
    if run_sync:
        ok, sync_output = sync_runner()
        if not ok:
            raise RuntimeError(sync_output or "Canonical sync failed after CategoryOverrides write-back.")

    return CategoryOverrideResult(
        sku=sku_norm,
        category=str(category).strip(),
        subcategory=str(subcategory).strip(),
        confidence=str(confidence).strip() or "manual",
        workbook_path=str(workbook_path),
        backup_path=str(backup_path) if backup_path else None,
        sync_output=sync_output,
    )
