from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.excel_sync import MASTER_PATH, backup_workbook, run_canonical_sync

SHEET_CATALOG = "Catalog"
SHEET_SALES = "DailySales"


@dataclass
class SaleWriteResult:
    sku: str
    product: str
    quantity: int
    unit_price: float
    total: float
    workbook_path: str
    backup_path: str | None = None
    sync_output: str = ""


def _norm_sku(value: object) -> str:
    if value is None:
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    try:
        return str(int(float(raw))).zfill(5)
    except ValueError:
        return raw


def _require_sheet(wb: openpyxl.Workbook, name: str) -> Worksheet:
    if name not in wb.sheetnames:
        raise ValueError(f"Required sheet missing from workbook: {name}")
    return wb[name]


def _catalog_lookup(ws: Worksheet) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = _norm_sku(row[0] if row else None)
        if sku:
            lookup[sku] = str(row[1] or "").strip()
    return lookup


def append_sale_to_excel(
    sale_date: date | datetime,
    sku: str,
    product: str,
    quantity: int,
    unit_price: float,
    payment_method: str,
    source: str = "manual",
    workbook_path: Path = MASTER_PATH,
    run_sync: bool = True,
    create_backup: bool = True,
    sync_runner=run_canonical_sync,
) -> SaleWriteResult:
    """Append a sale directly to the canonical DailySales worksheet, then sync."""
    sku_norm = _norm_sku(sku)
    if not sku_norm:
        raise ValueError("SKU is required.")
    if quantity <= 0:
        raise ValueError("Quantity must be greater than zero.")
    if unit_price < 0:
        raise ValueError("Unit_Price must be greater than or equal to zero.")

    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    backup_path = backup_workbook(workbook_path) if create_backup else None
    wb = openpyxl.load_workbook(workbook_path)
    catalog_ws = _require_sheet(wb, SHEET_CATALOG)
    sales_ws = _require_sheet(wb, SHEET_SALES)

    catalog = _catalog_lookup(catalog_ws)
    if sku_norm not in catalog:
        raise ValueError(f"SKU not found in Catalog: {sku_norm}")

    expected_product = catalog[sku_norm]
    product_name = (product or expected_product).strip() or expected_product
    total = round(quantity * float(unit_price), 2)
    row = [
        sale_date.strftime("%Y-%m-%d"),
        sku_norm,
        product_name,
        int(quantity),
        round(float(unit_price), 2),
        total,
        str(payment_method).strip(),
        str(source).strip() or "manual",
    ]
    sales_ws.append(row)
    wb.save(workbook_path)

    sync_output = ""
    if run_sync:
        ok, sync_output = sync_runner()
        if not ok:
            raise RuntimeError(sync_output or "Canonical sync failed after DailySales write-back.")

    return SaleWriteResult(
        sku=sku_norm,
        product=product_name,
        quantity=int(quantity),
        unit_price=round(float(unit_price), 2),
        total=total,
        workbook_path=str(workbook_path),
        backup_path=str(backup_path) if backup_path else None,
        sync_output=sync_output,
    )
