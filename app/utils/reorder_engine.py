"""
FulôFiló — Smart Reorder Alert Engine
=======================================
Thin wrapper over core.procurement.lead_time (per-SKU lead time + dynamic buffer).
"""

from __future__ import annotations

import os
import subprocess
from pathlib import Path

import pandas as pd

from core.procurement.lead_time import (
    COVERAGE_DAYS,
    DEFAULT_BUFFER_DAYS,
    DEFAULT_LEAD_TIME_DAYS,
    get_alerts,
    get_reorder_df,
    urgency_label,
)

ROOT = Path(__file__).resolve().parent.parent.parent

# Backward-compatible constants
SALES_PERIOD_DAYS = 61
ROLLING_VELOCITY_DAYS = 14
LEAD_TIME_DAYS = DEFAULT_LEAD_TIME_DAYS
BUFFER_DAYS = DEFAULT_BUFFER_DAYS
ALERT_THRESHOLD = LEAD_TIME_DAYS + BUFFER_DAYS


def export_excel(conn) -> Path | None:
    """Generate data/outputs/alertas_reposicao.xlsx."""
    df = get_reorder_df(conn)
    if df.empty:
        return None

    out_dir = ROOT / "data" / "outputs"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "alertas_reposicao.xlsx"

    display = df.copy()
    display["days_remaining"] = display["days_remaining"].astype(int)
    display["suggested_qty"] = display["suggested_qty"].astype(int)
    display["daily_rate"] = display["daily_rate"].round(2)
    display["urgency"] = display.apply(
        lambda r: urgency_label(float(r["days_remaining"]), float(r["lead_time"])),
        axis=1,
    )

    display = display.rename(columns={
        "product": "Produto",
        "category": "Categoria",
        "supplier_name": "Fornecedor",
        "current_stock": "Estoque Atual",
        "qty_sold": "Vendido (Período)",
        "qty_14d": "Vendido (14d)",
        "daily_rate": "Venda/Dia",
        "days_remaining": "Dias Restantes",
        "suggested_qty": "Qtd Sugerida (45d)",
        "lead_time": "Lead Time (dias)",
        "buffer": "Buffer (dias)",
        "alert_threshold": "Limiar Alerta (dias)",
        "urgency": "Urgência",
    }).drop(columns=["slug", "sku", "margin_pct", "unit_profit", "abc_class", "unit_cost",
                     "stock_unknown", "supplier", "supplier_id", "moq", "case_pack", "demand_std"],
            errors="ignore")

    threshold = df["alert_threshold"].max() if "alert_threshold" in df.columns else ALERT_THRESHOLD
    alerts_display = display[display["Dias Restantes"] <= threshold].copy()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        alerts_display.to_excel(writer, sheet_name="⚠️ Reposição Urgente", index=False)
        display.to_excel(writer, sheet_name="📦 Todos os Produtos", index=False)
        _style_workbook(writer.book)

    return out_path


def _style_workbook(wb) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HEADER_BG = "0D1117"
    URGENTE_BG = "3D1515"
    ATENCAO_BG = "2E2200"
    OK_BG = "0D1F1A"
    TEXT_COLOR = "E2E8F0"
    ACCENT = "00D4FF"

    thin = Border(
        left=Side(style="thin", color="1A1A2E"),
        right=Side(style="thin", color="1A1A2E"),
        bottom=Side(style="thin", color="1A1A2E"),
    )

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True, color=ACCENT, size=10)
            cell.fill = PatternFill("solid", fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin

        urgency_col = None
        for idx, cell in enumerate(ws[1], 1):
            if str(cell.value) == "Urgência":
                urgency_col = idx
                break

        for row in ws.iter_rows(min_row=2):
            urgency_val = ""
            if urgency_col:
                urgency_val = str(ws.cell(row=row[0].row, column=urgency_col).value or "")
            bg = URGENTE_BG if "URGENTE" in urgency_val else (ATENCAO_BG if "ATENÇÃO" in urgency_val else OK_BG)
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(color=TEXT_COLOR, size=9)
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
        ws.freeze_panes = "A2"


def notify_macos(alerts_df: pd.DataFrame, *, title_prefix: str = "FulôFiló AI — Reposição") -> None:
    if alerts_df.empty:
        return
    is_cloud = bool(os.environ.get("STREAMLIT_SHARING_MODE") or os.environ.get("IS_STREAMLIT_CLOUD"))
    if is_cloud:
        return

    count = len(alerts_df)
    lead_col = alerts_df["lead_time"] if "lead_time" in alerts_df.columns else LEAD_TIME_DAYS
    urgentes = alerts_df[alerts_df["days_remaining"] <= lead_col]
    n_urgent = len(urgentes)
    po_hint = ""
    if "supplier_name" in alerts_df.columns:
        suppliers = alerts_df["supplier_name"].dropna().unique()[:3]
        if len(suppliers):
            po_hint = f" | Fornecedores: {', '.join(map(str, suppliers))}"

    title = (
        f"🔴 {n_urgent} produto(s) URGENTE(S) para repor!"
        if n_urgent > 0
        else f"⚠️ {count} produto(s) precisam de reposição"
    )
    top3 = alerts_df.head(3)["product"].tolist()
    body = ", ".join(top3) + (f" + {count - 3} mais" if count > 3 else "") + po_hint

    script = (
        f'display notification "{body}" '
        f'with title "{title}" '
        f'subtitle "{title_prefix}"'
    )
    try:
        subprocess.run(["osascript", "-e", script], capture_output=True, timeout=5)
    except Exception:
        pass
