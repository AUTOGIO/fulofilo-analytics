"""Write approved purchase orders to Excel master (canonical write path)."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from app.utils.excel_sync import backup_workbook

ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_XLSX = ROOT / "data" / "excel" / "FuloFilo_Master.xlsx"
SHEET = "PurchaseOrders"
HEADERS = [
    "po_id", "supplier_id", "supplier_name", "sku", "product", "qty",
    "unit_cost", "line_total", "status", "created_at", "notes",
]


@dataclass
class ApprovePOResult:
    po_id: str
    lines_written: int
    backup_path: str


def append_po_to_excel(
    po: dict,
    *,
    workbook_path: Path | None = None,
    status: str = "approved",
) -> ApprovePOResult:
    """Append PO lines to PurchaseOrders sheet in Excel master."""
    xlsx = workbook_path or DEFAULT_XLSX
    if not xlsx.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx}")

    backup = backup_workbook(workbook_path=xlsx)
    wb = openpyxl.load_workbook(xlsx)
    if SHEET not in wb.sheetnames:
        ws = wb.create_sheet(SHEET)
        ws.append(HEADERS)
    else:
        ws = wb[SHEET]
        if ws.max_row == 0 or str(ws.cell(1, 1).value or "") != "po_id":
            ws.append(HEADERS)

    created = po.get("created_at") or datetime.now(timezone.utc).isoformat()
    po_id = str(po.get("po_id", ""))
    supplier_id = str(po.get("supplier_id", ""))
    supplier_name = str(po.get("supplier_name", ""))
    lines = po.get("lines") or []
    written = 0

    for line in lines:
        qty = int(line.get("qty") or 0)
        unit_cost = float(line.get("unit_cost") or 0)
        ws.append([
            po_id,
            supplier_id,
            supplier_name,
            str(line.get("sku", "")),
            str(line.get("product", "")),
            qty,
            unit_cost,
            round(qty * unit_cost, 2),
            status,
            created,
            str(line.get("rationale", "")),
        ])
        written += 1

    wb.save(xlsx)
    return ApprovePOResult(po_id=po_id, lines_written=written, backup_path=str(backup))
