"""
FulôFiló Analytics Pro — Excel Report Builder
==============================================
Generates a professional 9-sheet .xlsx workbook from parquet data.
Target: Microsoft 365 Universal Binary on macOS M3.

Usage:
    python excel/build_report.py
    # OR
    from excel.build_report import build_report
    path = build_report()
"""

from __future__ import annotations

import datetime
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data" / "parquet"
EXCEL_DIR = ROOT / "excel"

# ── Brand Colors (openpyxl hex = no leading #) ───────────────────────────────
C_GREEN_DARK   = "2D6A4F"
C_GREEN_MID    = "52B788"
C_GREEN_LIGHT  = "D8F3DC"
C_YELLOW       = "F2C94C"
C_YELLOW_LIGHT = "FFF9DB"
C_RED          = "E74C3C"
C_RED_LIGHT    = "FDECEA"
C_ORANGE       = "F4A261"
C_BLUE_HDR     = "1B4F72"
C_GREY_LIGHT   = "F5F5F5"
C_WHITE        = "FFFFFF"
C_BLACK        = "1A1A1A"

# ── Thin border helper ────────────────────────────────────────────────────────
_THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MED  = Side(style="medium", color="888888")
MED_BORDER  = Border(left=_MED, right=_MED, top=_MED, bottom=_MED)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color=C_BLACK, italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name="Arial")


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Style helpers ─────────────────────────────────────────────────────────────
def style_header_row(ws, row: int, n_cols: int, bg=C_GREEN_DARK, fg=C_WHITE):
    """Apply dark header style to a full row."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font   = _font(bold=True, size=10, color=fg)
        cell.fill   = _fill(bg)
        cell.border = THIN_BORDER
        cell.alignment = _align(h="center")


def style_data_row(ws, row: int, n_cols: int, bg=C_WHITE):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = _font(size=10)
        cell.fill      = _fill(bg)
        cell.border    = THIN_BORDER
        cell.alignment = _align()


def set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def fmt_brl(ws, cell_range: str):
    """Apply Brazilian Real currency format to a range."""
    for row in ws[cell_range]:
        for c in row:
            c.number_format = 'R$ #,##0.00'


def fmt_pct(ws, cell_range: str):
    for row in ws[cell_range]:
        for c in row:
            c.number_format = '0.00%'


def fmt_date(ws, cell_range: str):
    for row in ws[cell_range]:
        for c in row:
            c.number_format = 'DD/MM/YYYY'


# ── Data loaders ──────────────────────────────────────────────────────────────
def _load(name: str) -> pl.DataFrame:
    path = DATA_DIR / f"{name}.parquet"
    if path.exists():
        return pl.read_parquet(path)
    return pl.DataFrame()


def _load_all() -> dict[str, pl.DataFrame]:
    return {
        "products":        _load("products"),
        "inventory":       _load("inventory"),
        "daily_sales":     _load("daily_sales"),
        "cashflow":        _load("cashflow"),
        "revenue_report":  _load("revenue_report"),
        "profit_report":   _load("profit_report"),
        "quantity_report": _load("quantity_report"),
    }


# ── KPI card writer ───────────────────────────────────────────────────────────
def _kpi_card(ws, start_row: int, start_col: int,
              title: str, value: str, subtitle: str = "",
              delta: str = "", positive: bool | None = None):
    """Write a KPI card spanning 2 cols × 4 rows."""
    r, c = start_row, start_col
    ws.merge_cells(start_row=r,   start_column=c, end_row=r,   end_column=c+1)
    ws.merge_cells(start_row=r+1, start_column=c, end_row=r+1, end_column=c+1)
    ws.merge_cells(start_row=r+2, start_column=c, end_row=r+2, end_column=c+1)
    ws.merge_cells(start_row=r+3, start_column=c, end_row=r+3, end_column=c+1)

    title_cell = ws.cell(r,   c, title)
    val_cell   = ws.cell(r+1, c, value)
    sub_cell   = ws.cell(r+2, c, subtitle)
    dlt_cell   = ws.cell(r+3, c, delta)

    title_cell.font = _font(bold=True, size=9, color=C_WHITE)
    title_cell.fill = _fill(C_GREEN_DARK)
    title_cell.alignment = _align(h="center")

    val_cell.font = _font(bold=True, size=16, color=C_BLACK)
    val_cell.fill = _fill(C_GREY_LIGHT)
    val_cell.alignment = _align(h="center")

    sub_cell.font = _font(size=8, color="666666", italic=True)
    sub_cell.fill = _fill(C_GREY_LIGHT)
    sub_cell.alignment = _align(h="center")

    delta_color = C_GREEN_DARK if positive else (C_RED if positive is False else "888888")
    dlt_cell.font = _font(bold=True, size=9, color=delta_color)
    dlt_cell.fill = _fill(C_GREY_LIGHT)
    dlt_cell.alignment = _align(h="center")

    for row_off in range(4):
        for col_off in range(2):
            ws.cell(r + row_off, c + col_off).border = THIN_BORDER


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Dashboard
# ══════════════════════════════════════════════════════════════════════════════
def build_dashboard(ws, data: dict[str, pl.DataFrame]):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"
    products = data["products"]

    total_rev   = float(products["revenue"].sum()) if not products.is_empty() else 0.0
    total_profit= float(products["profit"].sum())  if not products.is_empty() else 0.0
    margin_pct  = (total_profit / total_rev) if total_rev else 0.0
    top_cat     = (products.group_by("category")
                   .agg(pl.col("revenue").sum())
                   .sort("revenue", descending=True)
                   .row(0)[0]) if not products.is_empty() else "—"
    inv         = data["inventory"]
    inv_value   = float((inv["current_stock"].cast(pl.Float64) *
                         products.select("unit_cost")["unit_cost"]).sum()
                        ) if not products.is_empty() and not inv.is_empty() else 0.0
    n_skus      = products.shape[0]

    ws.row_dimensions[1].height = 40
    ws.merge_cells("A1:N1")
    hdr = ws["A1"]
    hdr.value     = "🌺  FulôFiló Analytics Pro — Dashboard Executivo"
    hdr.font      = _font(bold=True, size=18, color=C_WHITE)
    hdr.fill      = _fill(C_GREEN_DARK)
    hdr.alignment = _align(h="center", v="center")

    ws.merge_cells("A2:N2")
    sub = ws["A2"]
    sub.value     = f"Gerado em {datetime.date.today().strftime('%d/%m/%Y')}  •  {n_skus} SKUs ativos"
    sub.font      = _font(size=10, color="888888", italic=True)
    sub.fill      = _fill(C_GREY_LIGHT)
    sub.alignment = _align(h="center")
    # KPI cards — row 4, two cols each, gap col between cards
    cards = [
        ("💰 Receita Total (MTD)",  f"R$ {total_rev:,.2f}",  "Todas as vendas", "", None),
        ("📈 Margem Bruta",         f"{margin_pct:.1%}",     "Lucro / Receita",  "↑ vs meta 40%", margin_pct >= 0.40),
        ("🏆 Top Categoria",        top_cat,                  "Maior receita",   "", None),
        ("🏷️ Valor em Estoque",    f"R$ {inv_value:,.2f}",  f"{n_skus} SKUs",  "", None),
        ("📦 Total SKUs",           str(n_skus),             "Produtos ativos", "", None),
    ]
    col_starts = [1, 3, 5, 7, 9]  # cards at cols A, C, E, G, I
    for (title, val, sub_txt, delta, pos), start_col in zip(cards, col_starts):
        _kpi_card(ws, 4, start_col, title, val, sub_txt, delta, pos)

    # Top 10 products mini-table
    ws.row_dimensions[9].height = 20
    ws.merge_cells("A9:N9")
    sec = ws["A9"]
    sec.value     = "TOP 10 PRODUTOS POR RECEITA"
    sec.font      = _font(bold=True, size=11, color=C_WHITE)
    sec.fill      = _fill(C_BLUE_HDR)
    sec.alignment = _align(h="left")

    top10_headers = ["Rank","Produto","Categoria","Receita (R$)","Qtd Vendida","Margem %","Classe ABC"]
    for ci, h in enumerate(top10_headers, 1):
        c = ws.cell(10, ci, h)
        c.font = _font(bold=True, size=9, color=C_WHITE)
        c.fill = _fill(C_GREEN_DARK)
        c.border = THIN_BORDER
        c.alignment = _align(h="center")

    if not products.is_empty():
        top10 = products.sort("revenue", descending=True).head(10)
        abc_colors = {"A": C_GREEN_LIGHT, "B": C_YELLOW_LIGHT, "C": C_RED_LIGHT}
        for ri, row in enumerate(top10.iter_rows(named=True), 11):
            bg = abc_colors.get(row.get("abc_class", "C"), C_WHITE)
            row_data = [ri-10, row["full_name"], row["category"],
                        row["revenue"], row["qty_sold"],
                        (row["margin_pct"] or 0) / 100, row["abc_class"]]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(ri, ci, val)
                cell.fill   = _fill(bg)
                cell.border = THIN_BORDER
                cell.font   = _font(size=9)
                if ci == 4:
                    cell.number_format = 'R$ #,##0.00'
                    cell.alignment = _align(h="right")
                elif ci == 6:
                    cell.number_format = '0.00%'
                    cell.alignment = _align(h="center")
                else:
                    cell.alignment = _align(h="center" if ci in (1, 7) else "left")

    set_col_widths(ws, [6, 32, 18, 16, 14, 12, 10, 6, 6, 6, 6, 6, 6, 6])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — ABC Analysis
# ══════════════════════════════════════════════════════════════════════════════
def build_abc(ws, data: dict[str, pl.DataFrame]):
    ws.freeze_panes = "A2"
    products = data["products"]

    headers = ["Rank","Cód.","Produto","Categoria","Receita (R$)",
               "Receita %","Acumulado %","Classe ABC","Margem %","Qtd Vendida"]
    n_cols = len(headers)
    style_header_row(ws, 1, n_cols)
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)

    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    if not products.is_empty():
        total_rev = float(products["revenue"].sum())
        df = products.sort("revenue", descending=True).with_row_index("rank_idx")
        abc_fills = {"A": C_GREEN_LIGHT, "B": C_YELLOW_LIGHT, "C": C_RED_LIGHT}

        for ri, row in enumerate(df.iter_rows(named=True), 2):
            bg = abc_fills.get(row.get("abc_class", "C"), C_WHITE)
            rev_pct = (row["revenue"] / total_rev) if total_rev else 0
            cum_pct = (row["cum_pct"] / 100) if row.get("cum_pct") else 0
            margin  = (row["margin_pct"] or 0) / 100
            row_vals = [
                int(row["rank_idx"]) + 1, row["sku"], row["full_name"],
                row["category"], row["revenue"], rev_pct, cum_pct,
                row["abc_class"], margin, row["qty_sold"] or 0,
            ]
            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(ri, ci, val)
                cell.fill = _fill(bg); cell.border = THIN_BORDER
                cell.font = _font(size=9)
                if   ci == 5: cell.number_format = 'R$ #,##0.00'; cell.alignment = _align(h="right")
                elif ci in (6,7,9): cell.number_format = '0.00%'; cell.alignment = _align(h="center")
                elif ci == 1: cell.alignment = _align(h="center")
                else: cell.alignment = _align()

        # Bar chart — Top 20 revenue
        n_data = min(20, df.shape[0])
        chart = BarChart()
        chart.type, chart.grouping = "col", "clustered"
        chart.title = "Top 20 Produtos por Receita (R$)"
        chart.y_axis.title, chart.x_axis.title = "Receita (R$)", "Produto"
        chart.style, chart.height, chart.width = 10, 14, 28

        rev_ref = Reference(ws, min_col=5, min_row=2, max_row=n_data+1)
        cats    = Reference(ws, min_col=3, min_row=2, max_row=n_data+1)
        chart.add_data(rev_ref); chart.set_categories(cats)
        ws.add_chart(chart, f"L3")

    set_col_widths(ws, [6, 8, 32, 18, 16, 12, 12, 10, 12, 12])

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Margin Matrix
# ══════════════════════════════════════════════════════════════════════════════
def build_margin_matrix(ws, data: dict[str, pl.DataFrame]):
    ws.freeze_panes = "A2"
    products = data["products"]

    headers = ["Produto","Categoria","Custo (R$)","Preço (R$)","Margem R$",
               "Margem %","Volume","Receita Total","Margem Total"]
    n_cols = len(headers)
    style_header_row(ws, 1, n_cols)
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)

    if not products.is_empty():
        for ri, row in enumerate(products.sort("revenue", descending=True).iter_rows(named=True), 2):
            margin_pct_val = (row["margin_pct"] or 0) / 100
            bg = (C_RED_LIGHT   if margin_pct_val < 0.10 else
                  C_YELLOW_LIGHT if margin_pct_val < 0.25 else
                  "FFF3E0"       if margin_pct_val < 0.40 else C_GREEN_LIGHT)
            vals = [row["full_name"], row["category"],
                    row["unit_cost"], row["suggested_price"], row["unit_profit"],
                    margin_pct_val, row["qty_sold"] or 0,
                    row["revenue"], row["profit"]]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(ri, ci, val)
                cell.fill = _fill(bg); cell.border = THIN_BORDER; cell.font = _font(size=9)
                if ci in (3,4,5,8,9):
                    cell.number_format = 'R$ #,##0.00'; cell.alignment = _align(h="right")
                elif ci == 6:
                    cell.number_format = '0.00%'; cell.alignment = _align(h="center")
                elif ci == 7:
                    cell.alignment = _align(h="center")
                else: cell.alignment = _align()

        # Pivot: margin by category
        n_data = products.shape[0]
        pivot_start = n_data + 3
        ws.cell(pivot_start, 1, "RESUMO POR CATEGORIA").font = _font(bold=True, size=10, color=C_WHITE)
        ws.cell(pivot_start, 1).fill = _fill(C_BLUE_HDR)
        ws.merge_cells(start_row=pivot_start, start_column=1, end_row=pivot_start, end_column=4)

        pivot_hdr = ["Categoria","Receita Total","Margem Total","Margem % Média"]
        style_header_row(ws, pivot_start+1, 4)
        for ci, h in enumerate(pivot_hdr, 1):
            ws.cell(pivot_start+1, ci, h)

        cat_pivot = (products.group_by("category")
                     .agg([pl.col("revenue").sum().alias("rev"),
                           pl.col("profit").sum().alias("profit"),
                           pl.col("margin_pct").mean().alias("avg_margin")])
                     .sort("rev", descending=True))
        for ri, row in enumerate(cat_pivot.iter_rows(named=True), pivot_start+2):
            vals = [row["category"], row["rev"], row["profit"], row["avg_margin"]/100]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(ri, ci, val)
                cell.border = THIN_BORDER; cell.font = _font(size=9)
                if ci in (2,3): cell.number_format = 'R$ #,##0.00'; cell.alignment=_align(h="right")
                elif ci == 4: cell.number_format = '0.00%'; cell.alignment=_align(h="center")

    set_col_widths(ws, [32, 18, 12, 12, 12, 10, 10, 14, 14])

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Inventory
# ══════════════════════════════════════════════════════════════════════════════
def build_inventory(ws, data: dict[str, pl.DataFrame]):
    ws.freeze_panes = "A2"
    inv = data["inventory"]
    products = data["products"]

    headers = ["SKU","Produto","Categoria","Estoque Atual","Estoque Mín",
               "Ponto Reposição","Status","Dias de Estoque","Valor (R$)"]
    n_cols = len(headers)
    style_header_row(ws, 1, n_cols)
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    if not inv.is_empty() and not products.is_empty():
        cost_map = dict(zip(products["sku"].to_list(), products["unit_cost"].to_list()))
        qty_map  = dict(zip(products["sku"].to_list(), (products["qty_sold"].fill_null(0) / 365).to_list()))

        status_fills = {
            "🔴 CRÍTICO": C_RED_LIGHT, "🟡 BAIXO": C_YELLOW_LIGHT,
            "🟢 OK": C_GREEN_LIGHT,   "⬆️ EXCESSO": "E8F4FD",
        }

        for ri, row in enumerate(inv.iter_rows(named=True), 2):
            qty  = row["current_stock"] or 0
            mins = row["min_stock"] or 0
            rpnt = row["reorder_qty"] or 0
            cost = cost_map.get(row["sku"], 0) or 0
            daily_sales = qty_map.get(row["sku"], 0) or 0
            days_stock  = int(qty / daily_sales) if daily_sales > 0 else 999
            value = qty * cost

            if   qty <= mins * 0.5: status = "🔴 CRÍTICO"
            elif qty <= mins:       status = "🟡 BAIXO"
            elif qty > rpnt * 2:   status = "⬆️ EXCESSO"
            else:                   status = "🟢 OK"

            bg = status_fills.get(status, C_WHITE)
            vals = [row["sku"], row["product"], row["category"],
                    qty, mins, rpnt, status, days_stock, value]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(ri, ci, val)
                cell.fill = _fill(bg); cell.border = THIN_BORDER; cell.font = _font(size=9)
                if ci == 9: cell.number_format = 'R$ #,##0.00'; cell.alignment = _align(h="right")
                elif ci in (4,5,6,8): cell.alignment = _align(h="center")
                elif ci == 7: cell.alignment = _align(h="center")
                else: cell.alignment = _align()

        # Summary row
        n_rows = inv.shape[0]
        sr = n_rows + 2
        ws.cell(sr, 1, "TOTAIS").font = _font(bold=True, size=10)
        ws.cell(sr, 1).fill = _fill(C_GREY_LIGHT)
        ws.cell(sr, 9, f"=SUM(I2:I{n_rows+1})").number_format = 'R$ #,##0.00'
        ws.cell(sr, 9).font = _font(bold=True)
        ws.cell(sr, 9).fill = _fill(C_GREY_LIGHT)

    set_col_widths(ws, [8, 32, 18, 14, 12, 16, 14, 14, 14])

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Daily Ops  (handles empty daily_sales gracefully)
# ══════════════════════════════════════════════════════════════════════════════
def build_daily_ops(ws, data: dict[str, pl.DataFrame]):
    ws.freeze_panes = "A2"
    sales = data["daily_sales"]

    headers = ["Data","Receita (R$)","Transações","Ticket Médio (R$)",
               "Produto Top","Método Pagto","Entrada (R$)","Saída (R$)","Saldo (R$)"]
    n_cols = len(headers)
    style_header_row(ws, 1, n_cols)
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)

    if not sales.is_empty():
        # ── Normalize schema: handle both CSV-schema and parquet-schema ────────
        cols = sales.columns
        # Rename lowercase variants to canonical names
        rename_map = {}
        if "date" in cols and "Date" not in cols:    rename_map["date"]           = "Date"
        if "revenue" in cols and "Total" not in cols: rename_map["revenue"]        = "Total"
        if "slug" in cols and "Product" not in cols:  rename_map["slug"]           = "Product"
        if "payment" in cols and "Payment_Method" not in cols: rename_map["payment"] = "Payment_Method"
        if rename_map:
            sales = sales.rename(rename_map)
        # Ensure required columns exist with fallbacks
        if "Product" not in sales.columns:
            sales = sales.with_columns(pl.lit("—").alias("Product"))
        if "Payment_Method" not in sales.columns:
            sales = sales.with_columns(pl.lit("—").alias("Payment_Method"))
        if "Total" not in sales.columns and "qty" in sales.columns:
            sales = sales.with_columns((pl.col("qty").cast(pl.Float64)).alias("Total"))

        daily = (sales.with_columns(pl.col("Date").cast(pl.Utf8))
                 .group_by("Date")
                 .agg([pl.col("Total").sum().alias("revenue"),
                       pl.col("Total").count().alias("transactions"),
                       pl.col("Total").mean().alias("avg_ticket"),
                       pl.col("Product").first().alias("top_product"),
                       pl.col("Payment_Method").first().alias("pay_method")])
                 .sort("Date"))
        for ri, row in enumerate(daily.iter_rows(named=True), 2):
            rev = row["revenue"] or 0
            vals = [row["Date"], rev, row["transactions"], row["avg_ticket"] or 0,
                    row["top_product"], row["pay_method"], rev, 0, rev]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(ri, ci, val)
                cell.border = THIN_BORDER; cell.font = _font(size=9)
                bg = C_GREY_LIGHT if ri % 2 == 0 else C_WHITE
                cell.fill = _fill(bg)
                if ci in (2,4,7,8,9): cell.number_format='R$ #,##0.00'; cell.alignment=_align(h="right")
                else: cell.alignment = _align()
    else:
        ws.cell(3, 1, "⚠️  Sem dados de vendas diárias. Execute etl/ingest_eleve.py para importar.")
        ws.cell(3, 1).font = _font(italic=True, color="888888", size=10)
        ws.merge_cells("A3:I3")

    set_col_widths(ws, [14, 16, 14, 16, 28, 18, 16, 16, 16])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Cashflow
# ══════════════════════════════════════════════════════════════════════════════
def build_cashflow(ws, data: dict[str, pl.DataFrame]):
    ws.freeze_panes = "A2"
    cf = data["cashflow"]

    headers = ["Mês","Saldo Inicial","Entradas (R$)","Saídas (R$)",
               "Fluxo Líquido","Saldo Final","Runway (meses)"]
    n_cols = len(headers)
    style_header_row(ws, 1, n_cols)
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)

    if not cf.is_empty():
        monthly = (cf.with_columns(pl.col("Date").str.slice(0,7).alias("month"))
                   .group_by("month")
                   .agg([pl.col("Amount").filter(pl.col("Type")=="Receita").sum().alias("in_"),
                         pl.col("Amount").filter(pl.col("Type")=="Despesa").sum().alias("out_")])
                   .sort("month"))
        balance = 0.0
        for ri, row in enumerate(monthly.iter_rows(named=True), 2):
            net = (row["in_"] or 0) - (row["out_"] or 0)
            bal_final = balance + net
            runway = round(bal_final / (row["out_"] or 1), 1) if (row["out_"] or 0) > 0 else 99.0
            bg = C_RED_LIGHT if net < 0 else C_WHITE
            vals = [row["month"], balance, row["in_"] or 0,
                    row["out_"] or 0, net, bal_final, runway]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(ri, ci, val)
                cell.fill=_fill(bg); cell.border=THIN_BORDER; cell.font=_font(size=9)
                if ci in (2,3,4,5,6): cell.number_format='R$ #,##0.00'; cell.alignment=_align(h="right")
                elif ci == 7: cell.number_format='0.0'; cell.alignment=_align(h="center")
                else: cell.alignment=_align()
            balance = bal_final
    else:
        ws.cell(3, 1, "⚠️  Sem dados de fluxo de caixa. Execute etl/ingest_eleve.py para importar.")
        ws.cell(3, 1).font = _font(italic=True, color="888888", size=10)
        ws.merge_cells("A3:G3")

    set_col_widths(ws, [14, 16, 16, 16, 16, 16, 14])

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — Products Catalog
# ══════════════════════════════════════════════════════════════════════════════
def build_products_catalog(ws, data: dict[str, pl.DataFrame]):
    ws.freeze_panes = "A2"
    products = data["products"]

    headers = ["SKU","Nome Completo","Categoria","Custo (R$)",
               "Preço Sugerido (R$)","Margem %","Qtd Vendida","Receita (R$)","Classe ABC"]
    n_cols = len(headers)
    style_header_row(ws, 1, n_cols)
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    if not products.is_empty():
        for ri, row in enumerate(products.sort("full_name").iter_rows(named=True), 2):
            bg = C_GREY_LIGHT if ri % 2 == 0 else C_WHITE
            vals = [row["sku"], row["full_name"], row["category"],
                    row["unit_cost"], row["suggested_price"],
                    (row["margin_pct"] or 0)/100, row["qty_sold"] or 0,
                    row["revenue"], row["abc_class"]]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(ri, ci, val)
                cell.fill=_fill(bg); cell.border=THIN_BORDER; cell.font=_font(size=9)
                if ci in (4,5,8): cell.number_format='R$ #,##0.00'; cell.alignment=_align(h="right")
                elif ci == 6: cell.number_format='0.00%'; cell.alignment=_align(h="center")
                elif ci in (1,9): cell.alignment=_align(h="center")
                else: cell.alignment=_align()

    set_col_widths(ws, [8, 34, 18, 14, 18, 12, 14, 16, 10])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — Product Categories
# ══════════════════════════════════════════════════════════════════════════════
def build_categories(ws, data: dict[str, pl.DataFrame]):
    ws.freeze_panes = "A2"
    products = data["products"]

    headers = ["Categoria","Total SKUs","Receita Total (R$)","Margem % Média","Classe ABC"]
    n_cols = len(headers)
    style_header_row(ws, 1, n_cols)
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)

    if not products.is_empty():
        cat_df = (products.group_by("category")
                  .agg([pl.col("sku").count().alias("n_skus"),
                        pl.col("revenue").sum().alias("total_rev"),
                        pl.col("margin_pct").mean().alias("avg_margin")])
                  .sort("total_rev", descending=True))

        total_rev = float(cat_df["total_rev"].sum())
        cum = 0.0
        for ri, row in enumerate(cat_df.iter_rows(named=True), 2):
            cum += (row["total_rev"] or 0)
            cum_pct = cum / total_rev if total_rev else 0
            abc = "A" if cum_pct <= 0.80 else ("B" if cum_pct <= 0.95 else "C")
            abc_fills = {"A": C_GREEN_LIGHT, "B": C_YELLOW_LIGHT, "C": C_RED_LIGHT}
            bg = abc_fills.get(abc, C_WHITE)

            vals = [row["category"], row["n_skus"],
                    row["total_rev"], (row["avg_margin"] or 0)/100, abc]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(ri, ci, val)
                cell.fill=_fill(bg); cell.border=THIN_BORDER; cell.font=_font(size=9)
                if ci == 3: cell.number_format='R$ #,##0.00'; cell.alignment=_align(h="right")
                elif ci == 4: cell.number_format='0.00%'; cell.alignment=_align(h="center")
                elif ci in (2,5): cell.alignment=_align(h="center")
                else: cell.alignment=_align()

    set_col_widths(ws, [24, 12, 18, 16, 12])

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 9 — Pivot Category × Month
# ══════════════════════════════════════════════════════════════════════════════
def build_pivot_cat_month(ws, data: dict[str, pl.DataFrame]):
    ws.freeze_panes = "B2"
    products = data["products"]
    sales    = data["daily_sales"]

    # If no transactional data, build a static category pivot from products
    if not products.is_empty():
        cats = (products.group_by("category")
                .agg(pl.col("revenue").sum().alias("revenue"))
                .sort("revenue", descending=True)
                ["category"].to_list())

        # Static columns: category totals (no real month data without daily_sales)
        ws.cell(1, 1, "Categoria").font = _font(bold=True, size=10, color=C_WHITE)
        ws.cell(1, 1).fill = _fill(C_BLUE_HDR)
        ws.cell(1, 1).alignment = _align(h="center")

        ws.cell(1, 2, "Receita Total (R$)").font = _font(bold=True, size=10, color=C_WHITE)
        ws.cell(1, 2).fill = _fill(C_BLUE_HDR)
        ws.cell(1, 2).alignment = _align(h="center")

        ws.cell(1, 3, "SKUs").font = _font(bold=True, size=10, color=C_WHITE)
        ws.cell(1, 3).fill = _fill(C_BLUE_HDR)
        ws.cell(1, 3).alignment = _align(h="center")

        ws.cell(1, 4, "Margem % Média").font = _font(bold=True, size=10, color=C_WHITE)
        ws.cell(1, 4).fill = _fill(C_BLUE_HDR)
        ws.cell(1, 4).alignment = _align(h="center")

        cat_df = (products.group_by("category")
                  .agg([pl.col("revenue").sum().alias("rev"),
                        pl.col("sku").count().alias("n"),
                        pl.col("margin_pct").mean().alias("m")])
                  .sort("rev", descending=True))

        color_scale = ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max",   end_color="52B788"
        )

        for ri, row in enumerate(cat_df.iter_rows(named=True), 2):
            bg = C_GREY_LIGHT if ri % 2 == 0 else C_WHITE
            vals = [row["category"], row["rev"], row["n"], (row["m"] or 0)/100]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(ri, ci, val)
                cell.fill=_fill(bg); cell.border=THIN_BORDER; cell.font=_font(size=9)
                if ci == 2: cell.number_format='R$ #,##0.00'; cell.alignment=_align(h="right")
                elif ci == 4: cell.number_format='0.00%'; cell.alignment=_align(h="center")
                elif ci == 3: cell.alignment=_align(h="center")
                else: cell.alignment=_align()

        # Color scale on revenue column
        n_cats = cat_df.shape[0]
        ws.conditional_formatting.add(f"B2:B{n_cats+1}", color_scale)

        # Grand total row
        tr = n_cats + 2
        ws.cell(tr, 1, "TOTAL GERAL").font = _font(bold=True, size=9)
        ws.cell(tr, 1).fill = _fill(C_GREY_LIGHT)
        ws.cell(tr, 2, f"=SUM(B2:B{n_cats+1})").number_format = 'R$ #,##0.00'
        ws.cell(tr, 2).font = _font(bold=True, size=9)
        ws.cell(tr, 2).fill = _fill(C_GREY_LIGHT)
        for ci in range(1, 5):
            ws.cell(tr, ci).border = THIN_BORDER

    set_col_widths(ws, [24, 18, 10, 16])

# ══════════════════════════════════════════════════════════════════════════════
# MAIN ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════
def build_report(output_path: Path | str | None = None) -> Path:
    """
    Build the complete FulôFiló Excel workbook.

    Args:
        output_path: Optional explicit output path. Defaults to
                     excel/FuloFilo_Report_YYYY-MM-DD.xlsx

    Returns:
        Path to the generated .xlsx file.
    """
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)

    if output_path is None:
        today = datetime.date.today().isoformat()
        output_path = EXCEL_DIR / f"FuloFilo_Report_{today}.xlsx"
    output_path = Path(output_path)

    print(f"[FulôFiló] Loading parquet data from {DATA_DIR} ...")
    data = _load_all()
    for name, df in data.items():
        print(f"  {name:20s}  {df.shape[0]:>5} rows  {list(df.schema.keys())[:4]}")

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    sheet_defs = [
        ("Dashboard",          build_dashboard),
        ("ABC Analysis",       build_abc),
        ("Margin Matrix",      build_margin_matrix),
        ("Inventory",          build_inventory),
        ("Daily Ops",          build_daily_ops),
        ("Cashflow",           build_cashflow),
        ("Products Catalog",   build_products_catalog),
        ("Product Categories", build_categories),
        ("Pivot Cat×Month",    build_pivot_cat_month),
    ]

    for sheet_name, builder_fn in sheet_defs:
        print(f"  → Building sheet: {sheet_name}")
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = C_GREEN_DARK
        # Page setup
        ws.page_setup.orientation = (
            "landscape" if sheet_name not in ("Products Catalog",) else "portrait"
        )
        ws.page_setup.paperSize = 9   # A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        try:
            builder_fn(ws, data)
        except Exception as exc:
            print(f"    ⚠ Error in {sheet_name}: {exc}")
            ws.cell(1, 1, f"ERROR: {exc}")

    wb.save(output_path)
    size_kb = output_path.stat().st_size // 1024
    print(f"\n✅ Report saved: {output_path}  ({size_kb} KB, {len(sheet_defs)} sheets)")
    return output_path


if __name__ == "__main__":
    path = build_report()
    print(path)
