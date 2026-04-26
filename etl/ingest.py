"""
FulôFiló — Smart Ingest
=======================
Single entry point for all data sources. Drop any file in data/input/ and run.

Detects file type by column signature — no hardcoded source names.

Supported formats
-----------------
CATALOG  (export_items.csv or any Excel/CSV with SKU + Name + Cost + Price[store])
  → rebuilds products.parquet + inventory.parquet

SALES    (item-sales-summary-YYYY-MM-DD-YYYY-MM-DD.csv or any CSV with
          Itens vendidos + Vendas líquidas + Custo das mercadorias)
  → appends daily_sales.parquet + cashflow.parquet
  → rebuilds analytics layer (ABC, revenue totals) on products.parquet

Usage
-----
    python etl/ingest.py data/input/export_items.csv
    python etl/ingest.py data/input/item-sales-summary-2026-05-01-2026-05-31.csv
    python etl/ingest.py data/input/                     # process all files in folder
    python etl/ingest.py --dry-run data/input/sales.csv
"""

from __future__ import annotations

import argparse
import csv
import datetime
import re
import sys
from pathlib import Path

import polars as pl

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE   = Path(__file__).resolve().parent.parent
INPUT  = BASE / "data" / "input"
OUT    = BASE / "data" / "parquet"

OUT.mkdir(parents=True, exist_ok=True)
INPUT.mkdir(parents=True, exist_ok=True)

# ── Column signatures for auto-detection ──────────────────────────────────────
# Catalog: must have all of these (case-insensitive)
CATALOG_REQUIRED = {"sku", "name", "category", "cost"}

# Sales: must have all of these
SALES_REQUIRED = {"sku", "itens vendidos", "vendas líquidas", "custo das mercadorias"}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _cols_lower(path: Path) -> set[str]:
    """Return lowercase column names from first row of CSV or Excel."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            cols = {str(c).lower().strip() for c in row if c is not None}
            wb.close()
            return cols
        wb.close()
        return set()
    else:
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, [])
        return {c.lower().strip() for c in header}


def _read_csv(path: Path) -> list[dict]:
    with open(path, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _read_excel(path: Path) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    result = [dict(zip(headers, row)) for row in rows[1:]]
    wb.close()
    return result


def _read_file(path: Path) -> list[dict]:
    if path.suffix.lower() in (".xlsx", ".xls"):
        return _read_excel(path)
    return _read_csv(path)


def _parse_float(v) -> float:
    if v is None:
        return 0.0
    s = str(v).replace(",", ".").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _write(df: pl.DataFrame, name: str, dry_run: bool) -> None:
    path = OUT / f"{name}.parquet"
    if dry_run:
        print(f"  [DRY-RUN] {name}.parquet ← {df.shape[0]} rows")
    else:
        df.write_parquet(path)
        print(f"  ✅ {name}.parquet — {df.shape[0]} rows")


def detect(path: Path) -> str:
    """Return 'catalog', 'sales', or 'unknown'."""
    cols = _cols_lower(path)
    if CATALOG_REQUIRED.issubset(cols):
        return "catalog"
    if SALES_REQUIRED.issubset(cols):
        return "sales"
    return "unknown"


# ══════════════════════════════════════════════════════════════════════════════
# CATALOG INGEST  →  products.parquet + inventory.parquet
# ══════════════════════════════════════════════════════════════════════════════

def _find_store_col(raw_headers: list[str], pattern: str) -> str | None:
    """Find a column matching pattern (case-insensitive), e.g. 'price ['."""
    pat = pattern.lower()
    for h in raw_headers:
        if h.lower().strip().startswith(pat):
            return h
    return None


def ingest_catalog(path: Path, dry_run: bool) -> None:
    print(f"\n── CATALOG: {path.name} ──────────────────────────────────")
    rows = _read_file(path)
    if not rows:
        print("  ⚠  File is empty.")
        return

    headers = list(rows[0].keys())

    # Resolve store-specific column names dynamically
    price_col    = _find_store_col(headers, "price [")    or "Price"
    in_stk_col   = _find_store_col(headers, "in stock [") or "In stock"
    low_stk_col  = _find_store_col(headers, "low stock [") or "Low stock"
    opt_stk_col  = _find_store_col(headers, "optimal stock [") or "Optimal stock"

    print(f"  Detected store columns: price='{price_col}' | "
          f"in_stock='{in_stk_col}' | low='{low_stk_col}' | opt='{opt_stk_col}'")

    # Preserve existing sales data if products.parquet already exists
    existing_sales: dict[str, dict] = {}
    prod_path = OUT / "products.parquet"
    if prod_path.exists():
        old = pl.read_parquet(prod_path)
        if "sku" in old.columns:
            for row in old.iter_rows(named=True):
                existing_sales[str(row["sku"])] = {
                    "qty_sold": float(row.get("qty_sold") or 0),
                    "revenue":  float(row.get("revenue")  or 0),
                    "profit":   float(row.get("profit")   or 0),
                }
        print(f"  Preserved sales data for {len(existing_sales)} existing SKUs")

    # Existing inventory stock counts
    existing_stock: dict[str, int] = {}
    inv_path = OUT / "inventory.parquet"
    if inv_path.exists():
        old_inv = pl.read_parquet(inv_path)
        if "sku" in old_inv.columns:
            for row in old_inv.iter_rows(named=True):
                existing_stock[str(row["sku"])] = int(row.get("current_stock") or 0)

    products_rows = []
    inventory_rows = []
    skipped = 0

    for row in rows:
        sku  = str(row.get("SKU") or "").strip()
        name = str(row.get("Name") or "").strip()
        cat  = str(row.get("Category") or "Outros").strip()

        if not sku or not name:
            skipped += 1
            continue

        cost  = _parse_float(row.get("Cost"))
        price = _parse_float(row.get(price_col))
        if price == 0:
            price = cost * 2  # fallback: 2× cost

        unit_profit = price - cost
        margin_pct  = round(unit_profit / price, 4) if price else 0.0
        min_stk     = int(_parse_float(row.get(low_stk_col)) or 10)
        reorder_qty = int(_parse_float(row.get(opt_stk_col)) or 30)

        # Sales fields: keep from existing parquet or zero
        sales = existing_sales.get(sku, {})
        qty_sold = sales.get("qty_sold", 0.0)
        revenue  = sales.get("revenue",  0.0)
        profit   = sales.get("profit",   0.0)

        products_rows.append({
            "sku":             sku,
            "full_name":       name,
            "category":        cat,
            "unit_cost":       cost,
            "suggested_price": price,
            "min_stock":       min_stk,
            "reorder_qty":     reorder_qty,
            "unit_profit":     round(unit_profit, 2),
            "margin_pct":      margin_pct,
            "qty_sold":        qty_sold,
            "revenue":         revenue,
            "profit":          profit,
        })

        # Stock: use file value → fallback to existing → fallback to 0
        in_stock_val = _parse_float(row.get(in_stk_col))
        current_stock = (int(in_stock_val) if in_stock_val > 0
                         else existing_stock.get(sku, 0))

        inventory_rows.append({
            "sku":           sku,
            "product":       name,
            "category":      cat,
            "current_stock": current_stock,
            "min_stock":     min_stk,
            "reorder_qty":   reorder_qty,
        })

    if skipped:
        print(f"  Skipped {skipped} rows with missing SKU or Name")

    # Build DataFrames
    prod_df = pl.DataFrame(products_rows, schema={
        "sku":             pl.String,
        "full_name":       pl.String,
        "category":        pl.String,
        "unit_cost":       pl.Float64,
        "suggested_price": pl.Float64,
        "min_stock":       pl.Int64,
        "reorder_qty":     pl.Int64,
        "unit_profit":     pl.Float64,
        "margin_pct":      pl.Float64,
        "qty_sold":        pl.Float64,
        "revenue":         pl.Float64,
        "profit":          pl.Float64,
    })

    # Compute ABC classification
    prod_df = _add_abc(prod_df)

    inv_df = pl.DataFrame(inventory_rows, schema={
        "sku":           pl.String,
        "product":       pl.String,
        "category":      pl.String,
        "current_stock": pl.Int64,
        "min_stock":     pl.Int64,
        "reorder_qty":   pl.Int64,
    })

    print(f"  {len(products_rows)} products | categories: "
          f"{prod_df['category'].n_unique()}")

    _write(prod_df, "products",  dry_run)
    _write(inv_df,  "inventory", dry_run)


# ══════════════════════════════════════════════════════════════════════════════
# SALES INGEST  →  daily_sales.parquet + cashflow.parquet + analytics rebuild
# ══════════════════════════════════════════════════════════════════════════════

def _date_range_from_filename(path: Path) -> tuple[str, str] | tuple[None, None]:
    """Extract start/end dates from filenames like item-sales-summary-2026-04-01-2026-04-30.csv"""
    m = re.search(r"(\d{4}-\d{2}-\d{2})-(\d{4}-\d{2}-\d{2})", path.stem)
    if m:
        return m.group(1), m.group(2)
    return None, None


def _working_days(start: str, end: str) -> list[str]:
    """Return Mon–Sat dates between start and end (inclusive)."""
    d = datetime.date.fromisoformat(start)
    e = datetime.date.fromisoformat(end)
    days = []
    while d <= e:
        if d.weekday() < 6:   # 0=Mon … 5=Sat
            days.append(d.isoformat())
        d += datetime.timedelta(days=1)
    return days or [start]   # at least one day even for same-day exports


def ingest_sales(path: Path, dry_run: bool) -> None:
    print(f"\n── SALES: {path.name} ──────────────────────────────────")

    rows = _read_file(path)
    if not rows:
        print("  ⚠  File is empty.")
        return

    # Resolve date range
    date_start, date_end = _date_range_from_filename(path)
    if not date_start:
        # Fallback: use today as single-day period
        today = datetime.date.today().isoformat()
        date_start = date_end = today
        print(f"  ⚠  No date range in filename — using today ({today})")
    else:
        print(f"  Period: {date_start} → {date_end}")

    work_days = _working_days(date_start, date_end)
    n_days    = len(work_days)
    source_id = path.stem   # used for deduplication

    # ── Build daily_sales entries ─────────────────────────────────────────────
    new_entries: list[dict] = []
    cashflow_rev  = 0.0
    cashflow_cost = 0.0

    for row in rows:
        item    = str(row.get("Item") or "").strip()
        qty     = _parse_float(row.get("Itens vendidos"))
        revenue = _parse_float(row.get("Vendas líquidas"))
        cost    = _parse_float(row.get("Custo das mercadorias") or 0)

        if not item or qty <= 0 or revenue <= 0:
            continue

        unit_price   = round(revenue / qty, 2) if qty else 0.0
        daily_qty    = qty     / n_days
        daily_rev    = revenue / n_days

        for day in work_days:
            new_entries.append({
                "Date":           day,
                "Product":        item,
                "Quantity":       round(daily_qty, 3),
                "Unit_Price":     unit_price,
                "Total":          round(daily_rev, 2),
                "Payment_Method": "Misto",
                "Source":         source_id,
            })

        cashflow_rev  += revenue
        cashflow_cost += cost

    if not new_entries:
        print("  ⚠  No valid rows found.")
        return

    print(f"  {len(rows)} products | {len(new_entries)} daily entries across "
          f"{n_days} days")

    # ── Merge into existing daily_sales.parquet (deduplicate by Source) ───────
    new_df = pl.DataFrame(new_entries, schema={
        "Date":           pl.String,
        "Product":        pl.String,
        "Quantity":       pl.Float64,
        "Unit_Price":     pl.Float64,
        "Total":          pl.Float64,
        "Payment_Method": pl.String,
        "Source":         pl.String,
    })

    ds_path = OUT / "daily_sales.parquet"
    if ds_path.exists():
        existing_ds = pl.read_parquet(ds_path)
        # Remove any rows from the same source (re-import = replace, not duplicate)
        if "Source" in existing_ds.columns:
            existing_ds = existing_ds.filter(pl.col("Source") != source_id)
        combined = pl.concat([existing_ds, new_df]).sort("Date")
    else:
        combined = new_df.sort("Date")

    _write(combined, "daily_sales", dry_run)

    # ── Merge into cashflow.parquet ───────────────────────────────────────────
    cf_rows = []
    if cashflow_rev > 0:
        cf_rows.append({
            "Date":           date_start,
            "Type":           "Receita",
            "Category":       "Vendas",
            "Description":    f"Vendas {date_start} → {date_end}",
            "Amount":         round(cashflow_rev, 2),
            "Payment_Method": "Misto",
        })
    if cashflow_cost > 0:
        cf_rows.append({
            "Date":           date_start,
            "Type":           "Despesa",
            "Category":       "CMV",
            "Description":    f"CMV {date_start} → {date_end}",
            "Amount":         round(cashflow_cost, 2),
            "Payment_Method": "Misto",
        })

    if cf_rows:
        cf_schema = {
            "Date":           pl.String,
            "Type":           pl.String,
            "Category":       pl.String,
            "Description":    pl.String,
            "Amount":         pl.Float64,
            "Payment_Method": pl.String,
        }
        new_cf = pl.DataFrame(cf_rows, schema=cf_schema)
        cf_path = OUT / "cashflow.parquet"
        if cf_path.exists():
            existing_cf = pl.read_parquet(cf_path)
            # Remove entries for overlapping period
            period_desc_prefix = f"Vendas {date_start}"
            if "Description" in existing_cf.columns:
                existing_cf = existing_cf.filter(
                    ~pl.col("Description").str.starts_with(period_desc_prefix)
                )
            combined_cf = pl.concat([existing_cf, new_cf]).sort("Date")
        else:
            combined_cf = new_cf.sort("Date")
        _write(combined_cf, "cashflow", dry_run)

        net = cashflow_rev - cashflow_cost
        print(f"  Receita R${cashflow_rev:,.2f} | CMV R${cashflow_cost:,.2f} | "
              f"Lucro R${net:,.2f} ({net/cashflow_rev:.1%})")

    # ── Rebuild analytics layer ───────────────────────────────────────────────
    if not dry_run:
        _rebuild_analytics(dry_run)


# ══════════════════════════════════════════════════════════════════════════════
# ANALYTICS REBUILD  (called after every sales ingest)
# ══════════════════════════════════════════════════════════════════════════════

def _add_abc(df: pl.DataFrame) -> pl.DataFrame:
    """Add cum_pct and abc_class columns based on revenue Pareto (80/15/5)."""
    total_rev = float(df["revenue"].sum())
    if total_rev == 0:
        return df.with_columns([
            pl.lit(0.0).alias("cum_pct"),
            pl.lit("C").alias("abc_class"),
        ])

    df = df.sort("revenue", descending=True)
    df = df.with_columns(
        (pl.col("revenue").cum_sum() / total_rev * 100).round(1).alias("cum_pct")
    )
    df = df.with_columns(
        pl.when(pl.col("cum_pct") <= 80).then(pl.lit("A"))
          .when(pl.col("cum_pct") <= 95).then(pl.lit("B"))
          .otherwise(pl.lit("C"))
          .alias("abc_class")
    )
    return df


def _rebuild_analytics(dry_run: bool) -> None:
    """Recompute revenue/qty_sold/profit/ABC on products.parquet from daily_sales."""
    prod_path = OUT / "products.parquet"
    ds_path   = OUT / "daily_sales.parquet"

    if not prod_path.exists():
        print("  ⚠  products.parquet not found — run catalog ingest first")
        return

    products  = pl.read_parquet(prod_path)

    if not ds_path.exists() or pl.read_parquet(ds_path).is_empty():
        # No sales yet — zero out and classify all C
        products = products.with_columns([
            pl.lit(0.0).alias("qty_sold"),
            pl.lit(0.0).alias("revenue"),
            pl.lit(0.0).alias("profit"),
            pl.lit(0.0).alias("cum_pct"),
            pl.lit("C").alias("abc_class"),
        ])
        _write(products, "products", dry_run)
        return

    daily_sales = pl.read_parquet(ds_path)

    # Aggregate sales by product name
    agg = (daily_sales
           .group_by("Product")
           .agg([
               pl.col("Quantity").sum().alias("qty_sold"),
               pl.col("Total").sum().alias("revenue"),
           ]))

    # Drop stale sales cols before joining fresh aggregates
    for col in ("qty_sold", "revenue", "profit"):
        if col in products.columns:
            products = products.drop(col)

    # Join on product name (sales use Name, products use full_name)
    products = (products
                .join(agg, left_on="full_name", right_on="Product", how="left")
                .with_columns([
                    pl.col("qty_sold").fill_null(0.0),
                    pl.col("revenue").fill_null(0.0),
                ])
                .with_columns(
                    # profit = revenue × catalog margin
                    (pl.col("revenue") * pl.col("margin_pct")).alias("profit")
                ))

    # Drop old analytics cols before re-adding
    for col in ("cum_pct", "abc_class"):
        if col in products.columns:
            products = products.drop(col)

    products = _add_abc(products)

    print(f"\n── Analytics rebuild ────────────────────────────────────")
    abc_summary = (products.group_by("abc_class")
                   .agg([pl.len().alias("n"), pl.col("revenue").sum().alias("rev")])
                   .sort("abc_class"))
    for row in abc_summary.iter_rows(named=True):
        print(f"  Class {row['abc_class']}: {row['n']:>3} products | "
              f"R$ {row['rev']:>10,.2f}")

    _write(products, "products", dry_run)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def process(path: Path, dry_run: bool) -> None:
    kind = detect(path)
    if kind == "catalog":
        ingest_catalog(path, dry_run)
    elif kind == "sales":
        ingest_sales(path, dry_run)
    else:
        print(f"  ⚠  {path.name} — unknown format. "
              f"Columns found: {_cols_lower(path)}")


def main() -> None:
    parser = argparse.ArgumentParser(description="FulôFiló — Smart Ingest")
    parser.add_argument("path", help="File or folder to ingest")
    parser.add_argument("--dry-run", action="store_true",
                        help="Validate without writing parquets")
    args = parser.parse_args()

    target = Path(args.path)
    dry    = args.dry_run

    print("=" * 60)
    print(f"FulôFiló Ingest  {'[DRY-RUN]' if dry else ''}")
    print("=" * 60)

    if target.is_dir():
        files = sorted(target.glob("*"))
        supported = [f for f in files
                     if f.suffix.lower() in (".csv", ".xlsx", ".xls")
                     and not f.name.startswith(".")]
        if not supported:
            print(f"  No CSV/Excel files found in {target}")
            sys.exit(0)
        for f in supported:
            process(f, dry)
    elif target.is_file():
        process(target, dry)
    else:
        print(f"  ⚠  Path not found: {target}")
        sys.exit(1)

    print("\n" + "=" * 60)
    print("✅ Ingest complete")
    print("=" * 60)


if __name__ == "__main__":
    main()
