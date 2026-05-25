"""
Export Canga SKUs 400–438 from FuloFilo_Master.xlsx (Catalog sheet) into a
canonical CSV for voice-memo transcript matching.

Usage:
    python etl/export_canga_voice_memo_catalog.py
    python etl/export_canga_voice_memo_catalog.py --master /path/to/FuloFilo_Master.xlsx
"""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_MASTER = ROOT / "data" / "excel" / "FuloFilo_Master.xlsx"
OUT_CSV = ROOT / "data" / "raw" / "voice_memo_transcripts" / "canga_catalog_400_438.csv"

SKU_MIN, SKU_MAX = 400, 438


def _strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def _slug(s: str) -> str:
    s = _strip_accents(s.lower())
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def _tipo_from_category(category: str) -> str | None:
    c = category or ""
    if "Elastano" in c:
        return "ELASTANO"
    if "Algodão" in c or "Algodao" in c:
        return "ALGODAO"
    return None


def _estampa_from_full_name(full_name: str) -> str:
    if "—" in full_name:
        return full_name.split("—", 1)[1].strip()
    return full_name.strip()


def _parse_sku(val) -> int | None:
    if val is None:
        return None
    try:
        return int(str(val).strip())
    except ValueError:
        return None


def main() -> None:
    ap = argparse.ArgumentParser(description="Export canga catalog snapshot for voice memos.")
    ap.add_argument("--master", type=Path, default=DEFAULT_MASTER, help="Path to FuloFilo_Master.xlsx")
    ap.add_argument("-o", "--output", type=Path, default=OUT_CSV, help="Output CSV path")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.master, read_only=True, data_only=True)
    ws = wb["Catalog"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header) if name}

    rows_out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = _parse_sku(row[idx["sku"]])
        if sku is None or not (SKU_MIN <= sku <= SKU_MAX):
            continue
        full_name = (row[idx["full_name"]] or "").strip()
        category = (row[idx["category"]] or "").strip()
        if not full_name:
            continue
        tipo = _tipo_from_category(category)
        if tipo is None:
            continue
        estampa = _estampa_from_full_name(full_name)
        estampa_flat = _strip_accents(estampa.lower())
        aliases = sorted(
            {
                estampa,
                estampa_flat,
                _slug(estampa),
                _slug(full_name),
            }
        )
        aliases_str = "|".join(a for a in aliases if a)
        rows_out.append(
            {
                "sku": sku,
                "tipo": tipo,
                "estampa_canonica": estampa,
                "full_name": full_name,
                "category": category,
                "aliases": aliases_str,
            }
        )

    wb.close()
    rows_out.sort(key=lambda r: r["sku"])

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["sku", "tipo", "estampa_canonica", "full_name", "category", "aliases"],
        )
        w.writeheader()
        w.writerows(rows_out)

    print(f"Wrote {len(rows_out)} rows to {args.output}")


if __name__ == "__main__":
    main()
