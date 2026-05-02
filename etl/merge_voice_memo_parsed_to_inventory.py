"""
Merge voice-memo parsed output (CSV or XLSX) into FuloFilo_Master.xlsx Inventory only.

Valid SKUs: 400–438. Quantities are summed per SKU across parsed rows, then
applied as additions to current_stock for existing rows, or new Inventory rows
are created from Catalog when missing.

Usage:
    python etl/merge_voice_memo_parsed_to_inventory.py --parsed-xlsx path.xlsx --dry-run
    python etl/merge_voice_memo_parsed_to_inventory.py --parsed-xlsx path.xlsx --dry-run --skip-needs-review
    python etl/merge_voice_memo_parsed_to_inventory.py --parsed-csv path.csv
"""

from __future__ import annotations

import argparse
import csv
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook.workbook import Workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from etl.parse_voice_memo_transcripts import _bool_cell

DEFAULT_MASTER = ROOT / "data" / "excel" / "FuloFilo_Master.xlsx"
BACKUPS_DIR = ROOT / "data" / "excel" / "backups"
SKU_MIN, SKU_MAX = 400, 438


def _parse_sku(val: Any) -> int | None:
    if val is None or val == "":
        return None
    try:
        s = int(float(str(val).strip()))
        if SKU_MIN <= s <= SKU_MAX:
            return s
    except (TypeError, ValueError):
        pass
    return None


def _parse_qty(val: Any) -> int | None:
    if val is None or val == "":
        return None
    try:
        q = int(float(val))
        if q < 0:
            return None
        return q
    except (TypeError, ValueError):
        return None


def load_parsed_csv(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            rows.append(dict(row))
    return rows


def load_parsed_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Parsed" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Parsed sheet missing in {path}")
    ws = wb["Parsed"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    idx = {h: i for i, h in enumerate(headers) if h}

    def col(row: tuple[Any, ...], name: str) -> Any:
        i = idx.get(name)
        if i is None:
            return None
        return row[i] if i < len(row) else None

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        rows.append(
            {
                "tipo_canga": col(row, "Tipo de Canga"),
                "estampa": col(row, "Estampa"),
                "quantidade": col(row, "Quantidade"),
                "sku": col(row, "sku"),
                "confidence": col(row, "confidence"),
                "needs_review": col(row, "needs_review"),
                "raw_line": col(row, "raw_line"),
                "source_file": col(row, "source_file"),
            }
        )
    wb.close()
    return rows


def aggregate_qty_by_sku(
    parsed: list[dict[str, Any]], *, skip_needs_review: bool = False
) -> dict[int, int]:
    totals: dict[int, int] = defaultdict(int)
    for row in parsed:
        if skip_needs_review and _bool_cell(row.get("needs_review")):
            continue
        sku = _parse_sku(row.get("sku"))
        if sku is None:
            continue
        q = _parse_qty(row.get("quantidade"))
        if q is None:
            continue
        totals[sku] += q
    return dict(totals)


def _header_map(ws: Any) -> dict[str, int]:
    headers = [c.value for c in ws[1]]
    return {str(h).strip(): i for i, h in enumerate(headers) if h is not None}


def catalog_by_sku(master: Path) -> dict[int, dict[str, Any]]:
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    ws = wb["Catalog"]
    h = _header_map(ws)
    out: dict[int, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if "sku" not in h:
            break
        sk = _parse_sku(vals[h["sku"]] if h["sku"] < len(vals) else None)
        if sk is None:
            continue
        rec: dict[str, Any] = {}
        for name, ci in h.items():
            rec[name] = vals[ci] if ci < len(vals) else None
        out[sk] = rec
    wb.close()
    return out


def plan_inventory_merge(master: Path, deltas: dict[int, int]) -> list[dict[str, Any]]:
    """Build a list of planned change dicts for dry-run / apply."""
    cat = catalog_by_sku(master)
    wb = openpyxl.load_workbook(master, read_only=True, data_only=True)
    inv = wb["Inventory"]
    inv_h = _header_map(inv)
    if "sku" not in inv_h:
        wb.close()
        raise ValueError("Inventory sheet missing 'sku' column")
    sku_ci = inv_h["sku"]
    cur_ci = inv_h.get("current_stock")

    existing: dict[int, tuple[int, Any]] = {}
    for r_i, row in enumerate(inv.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row)
        sk = _parse_sku(vals[sku_ci] if sku_ci < len(vals) else None)
        if sk is None:
            continue
        cur = vals[cur_ci] if cur_ci is not None and cur_ci < len(vals) else None
        existing[sk] = (r_i, cur)

    wb.close()

    changes: list[dict[str, Any]] = []
    for sku, delta in sorted(deltas.items()):
        if sku not in cat:
            changes.append(
                {"sku": sku, "action": "skip", "reason": "sku not in Catalog", "delta": delta}
            )
            continue
        if sku in existing:
            row_idx, stock_raw = existing[sku]
            try:
                old = int(float(stock_raw)) if stock_raw is not None and str(stock_raw).strip() != "" else 0
            except (TypeError, ValueError):
                old = 0
            new = old + delta
            changes.append(
                {
                    "sku": sku,
                    "action": "increment",
                    "row_1based": row_idx,
                    "old_current_stock": old,
                    "new_current_stock": new,
                    "delta": delta,
                }
            )
        else:
            rec = cat[sku]
            changes.append(
                {
                    "sku": sku,
                    "action": "append",
                    "product": rec.get("full_name"),
                    "category": rec.get("category"),
                    "current_stock": delta,
                    "min_stock": rec.get("min_stock"),
                    "reorder_qty": rec.get("reorder_qty"),
                    "delta": delta,
                }
            )
    return changes


def print_plan(changes: list[dict[str, Any]]) -> None:
    for c in changes:
        a = c["sku"]
        if c["action"] == "increment":
            print(
                f"  INCREMENT sku={a} row={c['row_1based']} "
                f"current_stock {c['old_current_stock']} + {c['delta']} -> {c['new_current_stock']}"
            )
        elif c["action"] == "append":
            print(
                f"  APPEND sku={a} current_stock={c['current_stock']} "
                f"product={c['product']!r} category={c['category']!r}"
            )
        else:
            print(f"  SKIP sku={a} delta={c['delta']} reason={c.get('reason')}")


def backup_master(master: Path, backups_dir: Path | None = None) -> Path:
    dest_dir = backups_dir if backups_dir is not None else BACKUPS_DIR
    dest_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = dest_dir / f"FuloFilo_Master_{ts}_before_voice_memo_inventory_merge.xlsx"
    shutil.copy2(master, dest)
    return dest


def apply_inventory_merge(master: Path, changes: list[dict[str, Any]]) -> None:
    wb: Workbook = openpyxl.load_workbook(master, data_only=False)
    inv = wb["Inventory"]
    inv_h = _header_map(inv)
    max_col = inv.max_column or len(inv[1])
    sku_ci = inv_h["sku"]
    cur_ci = inv_h["current_stock"]
    prod_ci = inv_h.get("product")
    cat_ci = inv_h.get("category")
    min_ci = inv_h.get("min_stock")
    reo_ci = inv_h.get("reorder_qty")
    sup_ci = inv_h.get("supplier")
    lead_ci = inv_h.get("lead_time_days")
    notes_ci = inv_h.get("notes")

    for c in changes:
        if c["action"] == "increment":
            r_i = c["row_1based"]
            inv.cell(row=r_i, column=cur_ci + 1, value=c["new_current_stock"])
        elif c["action"] == "append":
            new_row = [None] * max_col

            def set_col(ci: int | None, val: Any) -> None:
                if ci is not None and 0 <= ci < len(new_row):
                    new_row[ci] = val

            set_col(sku_ci, c["sku"])
            set_col(prod_ci, c.get("product"))
            set_col(cat_ci, c.get("category"))
            set_col(cur_ci, c.get("current_stock"))
            set_col(min_ci, c.get("min_stock"))
            set_col(reo_ci, c.get("reorder_qty"))
            set_col(sup_ci, None)
            set_col(lead_ci, 7)
            set_col(notes_ci, None)
            inv.append(new_row)

    wb.save(master)


def main() -> None:
    ap = argparse.ArgumentParser(description="Merge parsed voice memos into Master Inventory.")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--parsed-csv", type=Path, help="Parsed CSV from parse_voice_memo_transcripts.py")
    src.add_argument("--parsed-xlsx", type=Path, help="Parsed XLSX with sheet Parsed")
    ap.add_argument("--master", type=Path, default=DEFAULT_MASTER, help="Path to FuloFilo_Master.xlsx")
    ap.add_argument(
        "--backups-dir",
        type=Path,
        default=None,
        help=f"Backup folder (default: {BACKUPS_DIR})",
    )
    ap.add_argument("--dry-run", action="store_true", help="Print planned changes; do not write master")
    ap.add_argument(
        "--skip-needs-review",
        action="store_true",
        help="Only sum quantities from rows that do not need review (needs_review false/empty; "
        "excludes TRUE/1/yes/sim and Excel booleans/strings matching the parser)",
    )
    args = ap.parse_args()

    if args.parsed_csv:
        parsed = load_parsed_csv(args.parsed_csv)
        label = str(args.parsed_csv)
    else:
        parsed = load_parsed_xlsx(args.parsed_xlsx)
        label = str(args.parsed_xlsx)

    if not args.master.exists():
        print(f"Master workbook not found: {args.master}", file=sys.stderr)
        sys.exit(1)

    deltas = aggregate_qty_by_sku(parsed, skip_needs_review=args.skip_needs_review)
    changes = plan_inventory_merge(args.master, deltas)

    print(f"Parsed file: {label}")
    print(f"Parsed input rows: {len(parsed)}")
    if args.skip_needs_review:
        n_in = sum(1 for row in parsed if not _bool_cell(row.get("needs_review")))
        print(f"Rows included in aggregation (--skip-needs-review): {n_in}")
    print(f"SKUs with aggregated delta (400–438, valid qty): {len(deltas)}")
    print("Planned changes:")
    print_plan(changes)

    applicable = [c for c in changes if c["action"] in ("increment", "append")]
    if not applicable:
        print("Nothing to apply (no increment/append actions).")
        return

    if args.dry_run:
        print("[dry-run] No backup; master unchanged.")
        return

    backups_dir = args.backups_dir if args.backups_dir is not None else BACKUPS_DIR
    backup_path = backup_master(args.master, backups_dir)
    print(f"Backup: {backup_path}")
    apply_inventory_merge(args.master, changes)
    print(f"Updated Inventory in {args.master}")


if __name__ == "__main__":
    main()
