#!/usr/bin/env python3
"""
Export FuloFilo canga catalog from FuloFilo_Master.xlsx into CSV and JSON.

Input:
  /Users/giovannini_nuovo/Documents/GitHub/FuloFilo/data/excel/FuloFilo_Master.xlsx

Outputs:
  /Users/giovannini_nuovo/Documents/GitHub/FuloFilo/data/catalog/canga_catalog.csv
  /Users/giovannini_nuovo/Documents/GitHub/FuloFilo/data/catalog/canga_catalog.json

Design:
  - Stable pre-export catalog for Voice Memo parsing.
  - Filters SKU range 400–438.
  - Extracts tipo and estampa from workbook columns when available.
  - Falls back to product/name parsing when needed.
"""

from __future__ import annotations

import csv
import json
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl", file=sys.stderr)
    print("Install with: python3 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path("/Users/giovannini_nuovo/Documents/GitHub/FuloFilo")
WORKBOOK_PATH = REPO_ROOT / "data/excel/FuloFilo_Master.xlsx"
OUTPUT_DIR = REPO_ROOT / "data/catalog"
OUTPUT_CSV = OUTPUT_DIR / "canga_catalog.csv"
OUTPUT_JSON = OUTPUT_DIR / "canga_catalog.json"
SKU_MIN = 400
SKU_MAX = 438

PREFERRED_SHEET_NAMES = [
    "Catalog",
    "catalog",
    "Produtos",
    "products",
    "Product Catalog",
    "product_catalog",
]

SKU_HEADERS = {"sku", "codigo", "código", "id", "product_id"}
TIPO_HEADERS = {"tipo", "tipo de canga", "material", "categoria", "category"}
ESTAMPA_HEADERS = {"estampa", "estampa_canonica", "estampa canonical", "nome estampa"}
NAME_HEADERS = {"nome", "name", "produto", "product", "display name", "title", "descrição", "descricao"}


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = text.encode("ascii", "ignore").decode("utf-8")
    text = re.sub(r"\s+", " ", text)
    return text


def clean_display(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def slug_alias(value: str) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return text


def detect_tipo(*values: Any) -> str:
    combined = " ".join(normalize_text(v) for v in values if v is not None)
    if "elastano" in combined or "lycra" in combined or "canga_el" in combined:
        return "ELASTANO"
    if "algodao" in combined or "algodão" in combined or "canga_alg" in combined:
        return "ALGODAO"
    return "UNKNOWN"


def extract_estampa_from_name(name: str) -> str:
    display = clean_display(name)
    if "—" in display:
        return clean_display(display.split("—", 1)[1])
    if "-" in display:
        parts = display.split("-", 1)
        if len(parts) == 2 and "canga" in normalize_text(parts[0]):
            return clean_display(parts[1])

    norm = normalize_text(display)
    for prefix in [
        "canga elastano",
        "canga em elastano",
        "canga algodao",
        "canga em algodao",
        "canga",
    ]:
        if norm.startswith(prefix):
            stripped = display[len(prefix):].strip(" -—:")
            return clean_display(stripped) or display

    return display


def make_aliases(sku: str, tipo: str, estampa: str, raw_name: str = "") -> List[str]:
    candidates = {
        estampa,
        slug_alias(estampa),
        normalize_text(estampa),
        raw_name,
        slug_alias(raw_name),
        normalize_text(raw_name),
        sku,
    }

    # Avoid overly generic aliases.
    blocked = {"", "canga", "elastano", "algodao", "algodão", "unknown"}
    aliases = sorted({a for a in candidates if a and normalize_text(a) not in blocked})
    return aliases


def get_header_map(ws) -> Optional[Dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        raw_headers = [cell.value for cell in ws[row_idx]]
        headers = [normalize_text(h) for h in raw_headers]
        if any(h in SKU_HEADERS for h in headers):
            return {h: idx for idx, h in enumerate(headers) if h}
    return None


def find_col(header_map: Dict[str, int], candidates: set[str]) -> Optional[int]:
    for key, idx in header_map.items():
        if key in {normalize_text(c) for c in candidates}:
            return idx
    return None


def choose_sheet(wb):
    for name in PREFERRED_SHEET_NAMES:
        if name in wb.sheetnames:
            return wb[name]
    return wb[wb.sheetnames[0]]


def rows_from_sheet(ws) -> Iterable[Dict[str, Any]]:
    header_map = get_header_map(ws)
    if not header_map:
        raise RuntimeError(
            f"Could not find a header row with SKU column in sheet '{ws.title}'. "
            "Check workbook column names."
        )

    sku_col = find_col(header_map, SKU_HEADERS)
    tipo_col = find_col(header_map, TIPO_HEADERS)
    estampa_col = find_col(header_map, ESTAMPA_HEADERS)
    name_col = find_col(header_map, NAME_HEADERS)

    if sku_col is None:
        raise RuntimeError("SKU column not found.")

    # Header row is the first row where this header map was detected.
    header_row_index = None
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        headers = [normalize_text(cell.value) for cell in ws[row_idx]]
        if any(h in SKU_HEADERS for h in headers):
            header_row_index = row_idx
            break

    assert header_row_index is not None

    for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        raw_sku = row[sku_col] if sku_col < len(row) else None
        if raw_sku is None:
            continue

        sku_match = re.search(r"\d+", str(raw_sku))
        if not sku_match:
            continue

        sku_int = int(sku_match.group())
        if not (SKU_MIN <= sku_int <= SKU_MAX):
            continue

        sku = str(sku_int)
        raw_tipo = row[tipo_col] if tipo_col is not None and tipo_col < len(row) else ""
        raw_estampa = row[estampa_col] if estampa_col is not None and estampa_col < len(row) else ""
        raw_name = row[name_col] if name_col is not None and name_col < len(row) else ""

        tipo = detect_tipo(raw_tipo, raw_name, raw_estampa)
        estampa = clean_display(raw_estampa) or extract_estampa_from_name(clean_display(raw_name))

        if not estampa:
            estampa = f"SKU {sku}"

        yield {
            "sku": sku,
            "tipo": tipo,
            "estampa_canonical": estampa,
            "aliases": make_aliases(sku, tipo, estampa, clean_display(raw_name)),
        }


def write_outputs(records: List[Dict[str, Any]]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["sku", "tipo", "estampa_canonical", "aliases"])
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "sku": record["sku"],
                    "tipo": record["tipo"],
                    "estampa_canonical": record["estampa_canonical"],
                    "aliases": ";".join(record.get("aliases", [])),
                }
            )

    with OUTPUT_JSON.open("w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
        f.write("\n")


def main() -> int:
    if not WORKBOOK_PATH.exists():
        print(f"Workbook not found: {WORKBOOK_PATH}", file=sys.stderr)
        return 2

    wb = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    ws = choose_sheet(wb)

    records = list(rows_from_sheet(ws))
    records.sort(key=lambda r: int(r["sku"]))

    if not records:
        print("No SKU records found in range 400–438.", file=sys.stderr)
        return 3

    unknown_tipo = [r for r in records if r["tipo"] == "UNKNOWN"]
    write_outputs(records)

    print(f"Exported {len(records)} catalog rows.")
    print(f"CSV : {OUTPUT_CSV}")
    print(f"JSON: {OUTPUT_JSON}")

    if unknown_tipo:
        print("Warning: some records have UNKNOWN tipo. Review catalog manually:", file=sys.stderr)
        for r in unknown_tipo[:10]:
            print(f"  SKU {r['sku']}: {r['estampa_canonical']}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
